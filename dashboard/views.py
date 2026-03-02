from django.shortcuts import render
from core.decorators import login_required_json as login_required
from django.http import JsonResponse, HttpResponse
from django.db.models import Sum, Count, Q
from django.utils import timezone
from datetime import datetime, timedelta, date as dt_date
from calendar import monthrange
from core.decorators import role_required
from core.models import (
    Institut, Client, RendezVous, Paiement, Credit, PaiementCredit,
    ClotureCaisse, Employe, CarteCadeau, UtilisationCarteCadeau, ForfaitClient,
    Depense, Produit, VenteProduit,
)
import json


@login_required
@role_required(['patron'])
def index(request):
    """Dashboard principal du patron avec vue globale sur tous les instituts."""
    today = timezone.now().date()

    # Période sélectionnée (par défaut: aujourd'hui)
    periode = request.GET.get('periode', 'jour')

    if periode == 'jour':
        date_debut = today
        date_fin = today
    elif periode == 'semaine':
        date_debut = today - timedelta(days=today.weekday())
        date_fin = today
    elif periode == 'mois':
        date_debut = today.replace(day=1)
        date_fin = today
    elif periode == 'annee':
        date_debut = today.replace(month=1, day=1)
        date_fin = today
    elif periode == 'personnalisee':
        date_debut_str = request.GET.get('date_debut_custom', '')
        date_fin_str   = request.GET.get('date_fin_custom', '')
        try:
            date_debut = dt_date.fromisoformat(date_debut_str)
            date_fin   = dt_date.fromisoformat(date_fin_str)
            if date_fin < date_debut:
                date_fin = date_debut
        except ValueError:
            date_debut = today
            date_fin   = today
    else:
        date_debut = today.replace(day=1)
        date_fin = today

    # Récupérer les instituts
    instituts = Institut.objects.all()

    # Récupérer les dates clôturées pour Express dans la période
    # Les données Express ne doivent apparaître que si la caisse est clôturée
    express = Institut.objects.filter(code='express').first()
    dates_cloturees_express = []
    if express:
        dates_cloturees_express = list(
            ClotureCaisse.objects.filter(
                institut=express,
                date__range=[date_debut, date_fin],
                cloture=True
            ).values_list('date', flat=True)
        )

    # Stats globales - CA encaissé (paiements RDV validés)
    # Exclure carte_cadeau car déjà compté lors de la vente de la carte
    # Exclure forfait/offert (montant=0, pas d'argent encaissé)
    base_paiements = Paiement.objects.filter(
        rendez_vous__date__range=[date_debut, date_fin],
        rendez_vous__statut='valide'
    ).exclude(mode__in=['carte_cadeau', 'forfait', 'offert'])

    if express:
        # Non-Express
        ca_paiements_non_express = base_paiements.exclude(
            rendez_vous__institut=express
        ).aggregate(total=Sum('montant'))['total'] or 0
        # Express : uniquement dates clôturées
        ca_paiements_express = 0
        if dates_cloturees_express:
            ca_paiements_express = base_paiements.filter(
                rendez_vous__institut=express,
                rendez_vous__date__in=dates_cloturees_express
            ).aggregate(total=Sum('montant'))['total'] or 0
        ca_paiements_total = ca_paiements_non_express + ca_paiements_express
        # Garder la query complète pour le graphique par mode de paiement
        if dates_cloturees_express:
            paiements_query = base_paiements.exclude(
                rendez_vous__institut=express
            ) | base_paiements.filter(
                rendez_vous__institut=express,
                rendez_vous__date__in=dates_cloturees_express
            )
        else:
            paiements_query = base_paiements.exclude(rendez_vous__institut=express)
    else:
        ca_paiements_total = base_paiements.aggregate(total=Sum('montant'))['total'] or 0
        paiements_query = base_paiements

    # Cartes cadeaux vendues dans la période (ajoutées au CA total)
    base_cartes = CarteCadeau.objects.filter(
        date_achat__date__gte=date_debut,
        date_achat__date__lte=date_fin,
        statut__in=['active', 'soldee']
    )
    if express:
        ca_cartes_non_express = base_cartes.exclude(
            institut_achat=express
        ).aggregate(total=Sum('montant_initial'))['total'] or 0
        ca_cartes_express = 0
        if dates_cloturees_express:
            ca_cartes_express = base_cartes.filter(
                institut_achat=express,
                date_achat__date__in=dates_cloturees_express
            ).aggregate(total=Sum('montant_initial'))['total'] or 0
        ca_cartes_vendues_total = ca_cartes_non_express + ca_cartes_express
    else:
        ca_cartes_vendues_total = base_cartes.aggregate(total=Sum('montant_initial'))['total'] or 0

    # Crédits encaissés dans la période
    # Non-Express : filtre par plage de dates
    ca_credits_non_express = PaiementCredit.objects.filter(
        date__date__gte=date_debut,
        date__date__lte=date_fin
    ).exclude(credit__institut=express).aggregate(total=Sum('montant'))['total'] or 0 if express else \
        PaiementCredit.objects.filter(
            date__date__gte=date_debut,
            date__date__lte=date_fin
        ).aggregate(total=Sum('montant'))['total'] or 0

    # Express : uniquement les dates clôturées
    ca_credits_express = 0
    if express and dates_cloturees_express:
        ca_credits_express = PaiementCredit.objects.filter(
            credit__institut=express,
            date__date__in=dates_cloturees_express
        ).aggregate(total=Sum('montant'))['total'] or 0

    ca_credits_total = ca_credits_non_express + ca_credits_express

    # Forfaits vendus dans la période (montant encaissé à l'achat)
    # Les forfaits sont uniquement pour La Klinic (pas d'Express)
    ca_forfaits_total = ForfaitClient.objects.filter(
        date_achat__date__gte=date_debut,
        date_achat__date__lte=date_fin
    ).aggregate(total=Sum('montant_paye_initial'))['total'] or 0

    # Ventes de produits dans la période
    base_ventes_produits = VenteProduit.objects.filter(
        date__date__gte=date_debut,
        date__date__lte=date_fin
    )
    if express and not dates_cloturees_express:
        # Express sans clôture : exclure ses ventes
        ca_ventes_produits_total = base_ventes_produits.exclude(
            institut=express
        ).aggregate(total=Sum('montant_total'))['total'] or 0
    elif express and dates_cloturees_express:
        ca_ventes_produits_total = (
            base_ventes_produits.exclude(institut=express).aggregate(total=Sum('montant_total'))['total'] or 0
        ) + (
            base_ventes_produits.filter(
                institut=express, date__date__in=dates_cloturees_express
            ).aggregate(total=Sum('montant_total'))['total'] or 0
        )
    else:
        ca_ventes_produits_total = base_ventes_produits.aggregate(total=Sum('montant_total'))['total'] or 0

    ca_total = ca_paiements_total + ca_cartes_vendues_total + ca_credits_total + ca_forfaits_total + ca_ventes_produits_total

    # RDV de la période sélectionnée
    # Pour Express : n'inclure que les jours clôturés
    base_rdv = RendezVous.objects.filter(date__range=[date_debut, date_fin])
    base_rdv_valides = base_rdv.filter(statut='valide')

    if express:
        rdv_non_express = base_rdv.exclude(institut=express).count()
        rdv_valides_non_express = base_rdv_valides.exclude(institut=express).count()
        rdv_express = 0
        rdv_valides_express = 0
        if dates_cloturees_express:
            rdv_express = base_rdv.filter(institut=express, date__in=dates_cloturees_express).count()
            rdv_valides_express = base_rdv_valides.filter(institut=express, date__in=dates_cloturees_express).count()
        rdv_periode = rdv_non_express + rdv_express
        rdv_valides_periode = rdv_valides_non_express + rdv_valides_express
    else:
        rdv_periode = base_rdv.count()
        rdv_valides_periode = base_rdv_valides.count()

    # Crédits en cours
    credits_en_cours = Credit.objects.filter(solde=False).aggregate(
        total=Sum('montant_total') - Sum('montant_paye')
    )
    total_credits = credits_en_cours['total'] or 0

    # Nombre de clients avec dette
    clients_avec_dette = Credit.objects.filter(solde=False).values('client').distinct().count()

    # Clients actifs (ayant eu un RDV sur la période)
    # Pour Express : n'inclure que les jours clôturés
    base_clients_actifs = RendezVous.objects.filter(
        date__range=[date_debut, date_fin],
        statut='valide'
    )

    if express:
        clients_non_express = set(base_clients_actifs.exclude(
            institut=express
        ).values_list('client', flat=True))
        clients_express = set()
        if dates_cloturees_express:
            clients_express = set(base_clients_actifs.filter(
                institut=express, date__in=dates_cloturees_express
            ).values_list('client', flat=True))
        clients_actifs = len(clients_non_express | clients_express)
    else:
        clients_actifs = base_clients_actifs.values('client').distinct().count()

    # Stats par institut
    stats_par_institut = []
    for institut in instituts:
        # Pour Express : n'inclure que les dates clôturées
        if institut.code == 'express':
            if dates_cloturees_express:
                # CA de l'institut (seulement dates clôturées)
                ca_paiements = Paiement.objects.filter(
                    rendez_vous__institut=institut,
                    rendez_vous__date__in=dates_cloturees_express,
                    rendez_vous__statut='valide'
                ).exclude(mode__in=['carte_cadeau', 'forfait', 'offert']).aggregate(total=Sum('montant'))['total'] or 0

                # Cartes cadeaux vendues (dates clôturées)
                ca_cartes_vendues = CarteCadeau.objects.filter(
                    institut_achat=institut,
                    date_achat__date__in=dates_cloturees_express,
                    statut__in=['active', 'soldee']
                ).aggregate(total=Sum('montant_initial'))['total'] or 0

                # Crédits encaissés (dates clôturées)
                ca_credits = PaiementCredit.objects.filter(
                    credit__institut=institut,
                    date__date__in=dates_cloturees_express
                ).aggregate(total=Sum('montant'))['total'] or 0

                ca_forfaits = 0  # Pas de forfaits pour Express

                # Ventes produits (dates clôturées seulement)
                ca_ventes_produits = VenteProduit.objects.filter(
                    institut=institut,
                    date__date__in=dates_cloturees_express
                ).aggregate(total=Sum('montant_total'))['total'] or 0

                ca_institut = ca_paiements + ca_cartes_vendues + ca_credits + ca_ventes_produits

                # Nombre de RDV validés (seulement dates clôturées)
                rdv_count = RendezVous.objects.filter(
                    institut=institut,
                    date__in=dates_cloturees_express,
                    statut='valide'
                ).count()
            else:
                # Pas de dates clôturées : CA et RDV à 0
                ca_paiements = 0
                ca_cartes_vendues = 0
                ca_credits = 0
                ca_forfaits = 0
                ca_ventes_produits = 0
                ca_institut = 0
                rdv_count = 0
        else:
            # Autres instituts : comportement normal
            ca_paiements = Paiement.objects.filter(
                rendez_vous__institut=institut,
                rendez_vous__date__range=[date_debut, date_fin],
                rendez_vous__statut='valide'
            ).exclude(mode__in=['carte_cadeau', 'forfait', 'offert']).aggregate(total=Sum('montant'))['total'] or 0

            # Cartes cadeaux vendues dans cet institut
            ca_cartes_vendues = CarteCadeau.objects.filter(
                institut_achat=institut,
                date_achat__date__gte=date_debut,
                date_achat__date__lte=date_fin,
                statut__in=['active', 'soldee']
            ).aggregate(total=Sum('montant_initial'))['total'] or 0

            # Crédits encaissés dans cet institut
            ca_credits = PaiementCredit.objects.filter(
                credit__institut=institut,
                date__date__gte=date_debut,
                date__date__lte=date_fin
            ).aggregate(total=Sum('montant'))['total'] or 0

            # Forfaits vendus dans cet institut (encaissé à l'achat)
            ca_forfaits = ForfaitClient.objects.filter(
                institut=institut,
                date_achat__date__gte=date_debut,
                date_achat__date__lte=date_fin
            ).aggregate(total=Sum('montant_paye_initial'))['total'] or 0

            # Ventes produits dans cet institut
            ca_ventes_produits = VenteProduit.objects.filter(
                institut=institut,
                date__date__gte=date_debut,
                date__date__lte=date_fin
            ).aggregate(total=Sum('montant_total'))['total'] or 0

            ca_institut = ca_paiements + ca_cartes_vendues + ca_credits + ca_forfaits + ca_ventes_produits

            rdv_count = RendezVous.objects.filter(
                institut=institut,
                date__range=[date_debut, date_fin],
                statut='valide'
            ).count()

        # Clôtures du jour (peut y en avoir plusieurs)
        clotures_jour = ClotureCaisse.objects.filter(
            institut=institut,
            date=today,
            cloture=True
        )

        # Calculer le total cumulé de toutes les clôtures du jour
        total_jour = clotures_jour.aggregate(
            total=Sum('total_calcule')
        )['total'] or 0

        # Calculer les écarts du jour
        ecarts_jour = clotures_jour.aggregate(
            total_ecart=Sum('ecart'),
            total_especes_calcule=Sum('total_especes_calcule'),
            total_especes_reel=Sum('montant_reel_especes')
        )
        total_ecart = ecarts_jour['total_ecart'] or 0
        total_especes_calcule = ecarts_jour['total_especes_calcule'] or 0
        total_especes_reel = ecarts_jour['total_especes_reel'] or 0

        nb_clotures = clotures_jour.count()
        derniere_cloture = clotures_jour.order_by('-date_cloture').first()

        # Crédits en cours pour cet institut
        credits_institut = Credit.objects.filter(institut=institut, solde=False).aggregate(
            total=Sum('montant_total') - Sum('montant_paye')
        )
        credits_en_cours_institut = credits_institut['total'] or 0

        stats_par_institut.append({
            'institut': institut,
            'ca': ca_institut,
            'ca_paiements': ca_paiements,
            'ca_cartes_vendues': ca_cartes_vendues,
            'ca_credits': ca_credits,
            'ca_forfaits': ca_forfaits,
            'ca_ventes_produits': ca_ventes_produits,
            'credits_en_cours': credits_en_cours_institut,
            'rdv_count': rdv_count,
            'cloture': derniere_cloture,  # Pour compatibilité
            'clotures': clotures_jour,
            'nb_clotures': nb_clotures,
            'total_jour': total_jour,
            'total_ecart': total_ecart,
            'total_especes_calcule': total_especes_calcule,
            'total_especes_reel': total_especes_reel,
        })

    # Top 10 clients (par CA total)
    top_clients = Client.objects.annotate(
        total_ca=Sum('rendez_vous__paiements__montant', filter=Q(rendez_vous__statut='valide'))
    ).filter(total_ca__gt=0).order_by('-total_ca')[:10]

    # Clients avec dette (crédits non soldés)
    clients_dette = []
    credits_non_soldes = Credit.objects.filter(solde=False).select_related('client', 'institut')

    # Grouper par client
    dette_par_client = {}
    for credit in credits_non_soldes:
        client_id = credit.client.id
        if client_id not in dette_par_client:
            dette_par_client[client_id] = {
                'client': credit.client,
                'total_dette': 0,
                'nb_credits': 0
            }
        dette_par_client[client_id]['total_dette'] += credit.reste_a_payer
        dette_par_client[client_id]['nb_credits'] += 1

    clients_dette = sorted(dette_par_client.values(), key=lambda x: -x['total_dette'])[:10]

    # Derniers RDV validés
    derniers_rdv = RendezVous.objects.filter(
        statut='valide'
    ).select_related('client', 'employe', 'institut', 'prestation').order_by('-date', '-heure_debut')[:10]

    # Stats cartes cadeaux
    cartes_vendues_periode = CarteCadeau.objects.filter(
        date_achat__date__gte=date_debut,
        date_achat__date__lte=date_fin,
    ).aggregate(
        nombre=Count('id'),
        total=Sum('montant_initial'),
    )
    cartes_utilisees_periode = UtilisationCarteCadeau.objects.filter(
        date__date__gte=date_debut,
        date__date__lte=date_fin,
    ).aggregate(total=Sum('montant'))
    cartes_en_circulation = CarteCadeau.objects.filter(
        statut='active',
        solde__gt=0,
    ).aggregate(
        nombre=Count('id'),
        total=Sum('solde'),
    )
    cartes_recentes = CarteCadeau.objects.select_related(
        'acheteur', 'beneficiaire', 'institut_achat',
    ).order_by('-date_achat')[:10]

    # Stats forfaits (uniquement La Klinic)
    forfaits_vendus_periode = ForfaitClient.objects.filter(
        date_achat__date__gte=date_debut,
        date_achat__date__lte=date_fin,
    ).aggregate(
        nombre=Count('id'),
        total=Sum('prix_total'),
        seances_total=Sum('nombre_seances_total'),
    )
    forfaits_actifs = ForfaitClient.objects.filter(statut='actif').aggregate(
        nombre=Count('id'),
        seances_restantes=Sum('nombre_seances_total') - Sum('nombre_seances_utilisees'),
    )
    seances_effectuees_periode = RendezVous.objects.filter(
        date__range=[date_debut, date_fin],
        statut='valide',
        est_seance_forfait=True,
    ).count()
    forfaits_recents = ForfaitClient.objects.select_related(
        'client', 'prestation', 'institut',
    ).order_by('-date_achat')[:10]

    # CA global par moyen de paiement (pour le graphique)
    ca_par_paiement_query = paiements_query.values('mode').annotate(
        ca_total=Sum('montant')
    ).order_by('-ca_total')

    mode_display_map = {
        'especes': 'Espèces',
        'carte': 'Carte',
        'cheque': 'Chèque',
        'om': 'Orange Money',
        'wave': 'Wave',
        'carte_cadeau': 'Carte cadeau',
        'forfait': 'Forfait',
        'offert': 'Offert',
    }

    # Combiner paiements RDV + paiements crédits par mode
    ca_par_mode = {}
    for p in ca_par_paiement_query:
        if p['ca_total'] and p['ca_total'] > 0:
            mode = p['mode']
            ca_par_mode[mode] = ca_par_mode.get(mode, 0) + float(p['ca_total'])

    # Ajouter les paiements de crédits
    base_credits_paiements = PaiementCredit.objects.filter(
        date__date__gte=date_debut,
        date__date__lte=date_fin
    ).exclude(mode='carte_cadeau')
    if express:
        if dates_cloturees_express:
            credits_paiements_query = base_credits_paiements.filter(
                Q(~Q(credit__institut=express)) |
                Q(credit__institut=express, date__date__in=dates_cloturees_express)
            )
        else:
            credits_paiements_query = base_credits_paiements.exclude(credit__institut=express)
    else:
        credits_paiements_query = base_credits_paiements

    credits_par_mode = credits_paiements_query.values('mode').annotate(
        ca_total=Sum('montant')
    )
    for p in credits_par_mode:
        if p['ca_total'] and p['ca_total'] > 0:
            mode = p['mode']
            ca_par_mode[mode] = ca_par_mode.get(mode, 0) + float(p['ca_total'])

    # Ajouter les paiements des cartes cadeaux vendues
    cartes_paiement_query = base_cartes
    if express:
        if dates_cloturees_express:
            cartes_paiement_query = base_cartes.filter(
                Q(~Q(institut_achat=express)) |
                Q(institut_achat=express, date_achat__date__in=dates_cloturees_express)
            )
        else:
            cartes_paiement_query = base_cartes.exclude(institut_achat=express)
    for carte in cartes_paiement_query:
        mode1 = carte.mode_paiement_achat
        montant1 = carte.montant_paiement_1 if carte.montant_paiement_1 else carte.montant_initial
        ca_par_mode[mode1] = ca_par_mode.get(mode1, 0) + float(montant1)
        if carte.moyen_paiement_2 and carte.montant_paiement_2:
            ca_par_mode[carte.moyen_paiement_2] = ca_par_mode.get(carte.moyen_paiement_2, 0) + float(carte.montant_paiement_2)

    ca_par_paiement = sorted(
        [
            {'mode': mode_display_map.get(mode, mode), 'ca': ca}
            for mode, ca in ca_par_mode.items()
        ],
        key=lambda x: -x['ca']
    )

    # Clôtures de la période avec écarts
    clotures_periode = ClotureCaisse.objects.filter(
        date__range=[date_debut, date_fin],
        cloture=True
    ).select_related('institut', 'cloture_par__user').order_by('-date', '-date_cloture')

    # Total des écarts de la période
    total_ecart_periode = clotures_periode.aggregate(total=Sum('ecart'))['total'] or 0

    montrer_evolution_ca = not (
        periode == 'jour' or
        (periode == 'personnalisee' and date_debut == date_fin)
    )

    context = {
        'periode': periode,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'date_debut_custom': date_debut.isoformat() if periode == 'personnalisee' else '',
        'date_fin_custom':   date_fin.isoformat()   if periode == 'personnalisee' else '',
        'montrer_evolution_ca': montrer_evolution_ca,
        'ca_total': ca_total,
        'ca_paiements_total': ca_paiements_total,
        'ca_cartes_vendues_total': ca_cartes_vendues_total,
        'ca_credits_total': ca_credits_total,
        'ca_forfaits_total': ca_forfaits_total,
        'ca_ventes_produits_total': ca_ventes_produits_total,
        'rdv_periode': rdv_periode,
        'rdv_valides_periode': rdv_valides_periode,
        'total_credits': total_credits,
        'clients_avec_dette': clients_avec_dette,
        'clients_actifs': clients_actifs,
        'stats_par_institut': stats_par_institut,
        'top_clients': top_clients,
        'clients_dette': clients_dette,
        'derniers_rdv': derniers_rdv,
        'cartes_vendues_nombre': cartes_vendues_periode['nombre'] or 0,
        'cartes_vendues_montant': cartes_vendues_periode['total'] or 0,
        'cartes_utilisees_montant': cartes_utilisees_periode['total'] or 0,
        'cartes_circulation_nombre': cartes_en_circulation['nombre'] or 0,
        'cartes_circulation_montant': cartes_en_circulation['total'] or 0,
        'cartes_recentes': cartes_recentes,
        # Forfaits
        'forfaits_vendus_nombre': forfaits_vendus_periode['nombre'] or 0,
        'forfaits_vendus_montant': forfaits_vendus_periode['total'] or 0,
        'forfaits_seances_vendues': forfaits_vendus_periode['seances_total'] or 0,
        'forfaits_actifs_nombre': forfaits_actifs['nombre'] or 0,
        'forfaits_seances_restantes': forfaits_actifs['seances_restantes'] or 0,
        'seances_forfait_effectuees': seances_effectuees_periode,
        'forfaits_recents': forfaits_recents,
        # CA global par paiement
        'ca_par_paiement_json': json.dumps(ca_par_paiement),
        # Clôtures et écarts
        'clotures_periode': clotures_periode,
        'total_ecart_periode': total_ecart_periode,
    }

    # ── DÉBIT : Dépenses de la période ──
    depenses_qs = Depense.objects.filter(date__gte=date_debut, date__lte=date_fin)
    total_depenses = depenses_qs.aggregate(s=Sum('montant'))['s'] or 0

    depenses_par_cat = list(
        depenses_qs.values('categorie__nom').annotate(total=Sum('montant')).order_by('-total')[:6]
    )

    benefice = ca_total - total_depenses
    taux_charges = round(total_depenses / ca_total * 100, 1) if ca_total else 0

    # Alertes stock bas
    nb_alertes_stock = Produit.objects.filter(actif=True).count()
    # On utilise l'annotation pour éviter un loop Python — filtrage simple
    from django.db.models import F
    nb_alertes_stock = Produit.objects.filter(actif=True, stock_actuel__lte=F('stock_minimum')).count()

    context.update({
        'total_depenses': total_depenses,
        'depenses_par_cat': depenses_par_cat,
        'benefice': benefice,
        'taux_charges': taux_charges,
        'nb_alertes_stock': nb_alertes_stock,
    })

    # ── Onglet Bilan mensuel ──
    active_tab = request.GET.get('tab', 'dashboard')
    mois_param = request.GET.get('mois', '')
    bilan_institut_code = request.GET.get('bilan_institut', 'tous')

    if mois_param:
        try:
            from datetime import datetime as _dt
            bilan_mois = _dt.strptime(mois_param, '%Y-%m').date()
        except ValueError:
            bilan_mois = today.replace(day=1)
    else:
        bilan_mois = today.replace(day=1)

    bilan_fin_mois = (bilan_mois.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)

    bilan_paiements_qs = Paiement.objects.filter(
        date__date__gte=bilan_mois,
        date__date__lte=bilan_fin_mois,
    ).exclude(mode__in=['forfait', 'offert'])
    if bilan_institut_code != 'tous':
        bilan_paiements_qs = bilan_paiements_qs.filter(
            rendez_vous__employe__institut__code=bilan_institut_code
        )
    bilan_ca_rdv = bilan_paiements_qs.aggregate(s=Sum('montant'))['s'] or 0

    bilan_recettes_par_mode = {}
    for p in bilan_paiements_qs.values('mode').annotate(total=Sum('montant')).order_by('-total'):
        bilan_recettes_par_mode[p['mode']] = p['total']

    bilan_ventes_qs = VenteProduit.objects.filter(
        date__date__gte=bilan_mois,
        date__date__lte=bilan_fin_mois,
    )
    if bilan_institut_code != 'tous':
        bilan_ventes_qs = bilan_ventes_qs.filter(institut__code=bilan_institut_code)
    bilan_ca_ventes = bilan_ventes_qs.aggregate(s=Sum('montant_total'))['s'] or 0

    bilan_total_recettes = bilan_ca_rdv + bilan_ca_ventes

    bilan_dep_qs = Depense.objects.filter(date__gte=bilan_mois, date__lte=bilan_fin_mois)
    if bilan_institut_code != 'tous':
        bilan_dep_qs = bilan_dep_qs.filter(
            Q(institut__code=bilan_institut_code) | Q(institut__isnull=True)
        )
    bilan_total_depenses = bilan_dep_qs.aggregate(s=Sum('montant'))['s'] or 0
    bilan_depenses_par_cat = list(
        bilan_dep_qs.values('categorie__nom').annotate(total=Sum('montant')).order_by('-total')
    )
    bilan_benefice = bilan_total_recettes - bilan_total_depenses
    bilan_taux_charges = round((bilan_total_depenses / bilan_total_recettes * 100), 1) if bilan_total_recettes else 0

    try:
        nb_mois_bilan = int(request.GET.get('nb_mois', 6))
        if nb_mois_bilan not in (3, 6, 12):
            nb_mois_bilan = 6
    except (ValueError, TypeError):
        nb_mois_bilan = 6

    bilan_evolution = []
    for i in range(nb_mois_bilan - 1, -1, -1):
        month = bilan_mois.month - i
        year = bilan_mois.year
        while month <= 0:
            month += 12
            year -= 1
        m_debut = dt_date(year, month, 1)
        m_fin = (m_debut.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)

        rec_qs = Paiement.objects.filter(
            date__date__gte=m_debut, date__date__lte=m_fin,
        ).exclude(mode__in=['forfait', 'offert'])
        if bilan_institut_code != 'tous':
            rec_qs = rec_qs.filter(rendez_vous__employe__institut__code=bilan_institut_code)
        rec = rec_qs.aggregate(s=Sum('montant'))['s'] or 0

        vp_qs = VenteProduit.objects.filter(date__date__gte=m_debut, date__date__lte=m_fin)
        if bilan_institut_code != 'tous':
            vp_qs = vp_qs.filter(institut__code=bilan_institut_code)
        rec += vp_qs.aggregate(s=Sum('montant_total'))['s'] or 0

        dep_qs = Depense.objects.filter(date__gte=m_debut, date__lte=m_fin)
        if bilan_institut_code != 'tous':
            dep_qs = dep_qs.filter(Q(institut__code=bilan_institut_code) | Q(institut__isnull=True))
        dep = dep_qs.aggregate(s=Sum('montant'))['s'] or 0

        bilan_evolution.append({
            'label': m_debut.strftime('%b %Y'),
            'recettes': rec,
            'depenses': dep,
            'benefice': rec - dep,
        })

    bilan_evolution_json = json.dumps(bilan_evolution)

    context.update({
        'active_tab': active_tab,
        'bilan_mois': bilan_mois,
        'bilan_fin_mois': bilan_fin_mois,
        'bilan_institut_code': bilan_institut_code,
        'nb_mois_bilan': nb_mois_bilan,
        'bilan_ca_rdv': bilan_ca_rdv,
        'bilan_ca_ventes': bilan_ca_ventes,
        'bilan_total_recettes': bilan_total_recettes,
        'bilan_total_depenses': bilan_total_depenses,
        'bilan_depenses_par_cat': bilan_depenses_par_cat,
        'bilan_recettes_par_mode': bilan_recettes_par_mode,
        'bilan_benefice': bilan_benefice,
        'bilan_taux_charges': bilan_taux_charges,
        'bilan_evolution_json': bilan_evolution_json,
    })

    return render(request, 'dashboard/dashboard.html', context)


@login_required
@role_required(['patron'])
def api_stats_chart(request):
    """API pour les données des graphiques selon la période."""
    today = timezone.now().date()
    periode = request.GET.get('periode', 'mois')

    # Définir les dates et la granularité selon la période
    # IMPORTANT : mêmes plages que le dashboard index pour cohérence des totaux
    if periode == 'jour':
        date_debut = today
        date_fin = today
        labels = ['Aujourd\'hui']
        dates = [today]
    elif periode == 'semaine':
        date_debut = today - timedelta(days=today.weekday())
        date_fin = today
        nb_jours = (date_fin - date_debut).days
        dates = [date_debut + timedelta(days=i) for i in range(nb_jours + 1)]
        labels = [d.strftime('%d/%m') for d in dates]
    elif periode == 'mois':
        date_debut = today.replace(day=1)
        date_fin = today
        nb_jours = (date_fin - date_debut).days
        dates = [date_debut + timedelta(days=i) for i in range(nb_jours + 1)]
        labels = [d.strftime('%d/%m') for d in dates]
    else:  # annee
        date_debut = today.replace(month=1, day=1)
        date_fin = today
        labels = []
        dates = []
        for m in range(1, today.month + 1):
            labels.append(f'{m:02d}/{today.year}')
            dates.append((today.year, m))

    # Récupérer les dates clôturées pour Express
    express = Institut.objects.filter(code='express').first()
    dates_cloturees_express = set()
    if express:
        dates_cloturees_express = set(
            ClotureCaisse.objects.filter(
                institut=express,
                date__range=[date_debut, date_fin],
                cloture=True
            ).values_list('date', flat=True)
        )

    # Fonction utilitaire pour calculer le CA total d'une date ou période
    def calculer_ca_complet(d_debut, d_fin, institut_filter=None):
        """Calcule le CA complet (paiements + cartes + crédits + forfaits)."""
        # Paiements RDV
        paiements_qs = Paiement.objects.filter(
            rendez_vous__date__range=[d_debut, d_fin],
            rendez_vous__statut='valide'
        ).exclude(mode__in=['carte_cadeau', 'forfait', 'offert'])
        if institut_filter:
            paiements_qs = paiements_qs.filter(rendez_vous__institut=institut_filter)
            # Express : seulement dates clôturées
            if institut_filter.code == 'express' and dates_cloturees_express:
                paiements_qs = paiements_qs.filter(rendez_vous__date__in=dates_cloturees_express)
        elif express:
            if dates_cloturees_express:
                paiements_qs = paiements_qs.filter(
                    Q(~Q(rendez_vous__institut=express)) |
                    Q(rendez_vous__institut=express, rendez_vous__date__in=dates_cloturees_express)
                )
            else:
                paiements_qs = paiements_qs.exclude(rendez_vous__institut=express)
        ca_paiements = paiements_qs.aggregate(total=Sum('montant'))['total'] or 0

        # Cartes cadeaux vendues
        cartes_qs = CarteCadeau.objects.filter(
            statut__in=['active', 'soldee'],
            date_achat__date__gte=d_debut,
            date_achat__date__lte=d_fin
        )
        credits_qs = PaiementCredit.objects.filter(
            date__date__gte=d_debut,
            date__date__lte=d_fin
        )
        forfaits_qs = ForfaitClient.objects.filter(
            date_achat__date__gte=d_debut,
            date_achat__date__lte=d_fin
        )

        if institut_filter:
            cartes_qs = cartes_qs.filter(institut_achat=institut_filter)
            credits_qs = credits_qs.filter(credit__institut=institut_filter)
            forfaits_qs = forfaits_qs.filter(institut=institut_filter)

            if institut_filter.code == 'express':
                if dates_cloturees_express:
                    ca_cartes = cartes_qs.filter(date_achat__date__in=dates_cloturees_express).aggregate(total=Sum('montant_initial'))['total'] or 0
                    ca_credits = credits_qs.filter(date__date__in=dates_cloturees_express).aggregate(total=Sum('montant'))['total'] or 0
                else:
                    ca_cartes = 0
                    ca_credits = 0
                ca_forfaits = 0  # Pas de forfaits pour Express
            else:
                ca_cartes = cartes_qs.aggregate(total=Sum('montant_initial'))['total'] or 0
                ca_credits = credits_qs.aggregate(total=Sum('montant'))['total'] or 0
                ca_forfaits = forfaits_qs.aggregate(total=Sum('montant_paye_initial'))['total'] or 0
        elif express:
            if dates_cloturees_express:
                ca_cartes = (cartes_qs.exclude(institut_achat=express).aggregate(total=Sum('montant_initial'))['total'] or 0) + \
                    (cartes_qs.filter(institut_achat=express, date_achat__date__in=dates_cloturees_express).aggregate(total=Sum('montant_initial'))['total'] or 0)
                ca_credits = (credits_qs.exclude(credit__institut=express).aggregate(total=Sum('montant'))['total'] or 0) + \
                    (credits_qs.filter(credit__institut=express, date__date__in=dates_cloturees_express).aggregate(total=Sum('montant'))['total'] or 0)
            else:
                ca_cartes = cartes_qs.exclude(institut_achat=express).aggregate(total=Sum('montant_initial'))['total'] or 0
                ca_credits = credits_qs.exclude(credit__institut=express).aggregate(total=Sum('montant'))['total'] or 0
            ca_forfaits = forfaits_qs.aggregate(total=Sum('montant_paye_initial'))['total'] or 0
        else:
            ca_cartes = cartes_qs.aggregate(total=Sum('montant_initial'))['total'] or 0
            ca_credits = credits_qs.aggregate(total=Sum('montant'))['total'] or 0
            ca_forfaits = forfaits_qs.aggregate(total=Sum('montant_paye_initial'))['total'] or 0

        return ca_paiements + ca_cartes + ca_credits + ca_forfaits

    # Données CA évolution
    data_ca = []
    if periode == 'annee':
        for year, month in dates:
            last_day = monthrange(year, month)[1]
            if month == today.month and year == today.year:
                last_day = today.day
            d_debut_mois = dt_date(year, month, 1)
            d_fin_mois = dt_date(year, month, last_day)
            ca = calculer_ca_complet(d_debut_mois, d_fin_mois)
            data_ca.append({'date': f'{month:02d}/{year}', 'ca': ca})
    else:
        for i, date in enumerate(dates):
            ca = calculer_ca_complet(date, date)
            data_ca.append({'date': labels[i], 'ca': ca})

    # CA par institut
    instituts = Institut.objects.all()
    data_instituts = {}

    for institut in instituts:
        data_instituts[institut.nom] = []
        if periode == 'annee':
            for year, month in dates:
                last_day = monthrange(year, month)[1]
                if month == today.month and year == today.year:
                    last_day = today.day

                if institut.code == 'express' and not dates_cloturees_express:
                    ca = 0
                else:
                    d_debut_mois = dt_date(year, month, 1)
                    d_fin_mois = dt_date(year, month, last_day)
                    ca = calculer_ca_complet(d_debut_mois, d_fin_mois, institut)

                data_instituts[institut.nom].append(ca)
        else:
            for date in dates:
                if institut.code == 'express' and date not in dates_cloturees_express:
                    ca = 0
                else:
                    ca = calculer_ca_complet(date, date, institut)

                data_instituts[institut.nom].append(ca)

    return JsonResponse({
        'ca_evolution': data_ca,
        'ca_par_institut': data_instituts,
        'labels': labels,
        'periode': periode
    })


@login_required
@role_required(['patron'])
def api_stats_institut(request):
    """API pour les statistiques détaillées d'un institut (CA par employé + prestations)."""
    institut_code = request.GET.get('institut', 'palais')
    periode = request.GET.get('periode', 'mois')

    try:
        institut = Institut.objects.get(code=institut_code)
    except Institut.DoesNotExist:
        return JsonResponse({'error': 'Institut non trouvé'}, status=404)

    today = timezone.now().date()

    # Définir les dates selon la période — identique à dashboard/views.py index()
    if periode == 'jour':
        date_debut = today
        date_fin = today
    elif periode == 'semaine':
        date_debut = today - timedelta(days=today.weekday())
        date_fin = today
    elif periode == 'mois':
        date_debut = today.replace(day=1)
        date_fin = today
    elif periode == 'annee':
        date_debut = today.replace(month=1, day=1)
        date_fin = today
    elif periode == 'personnalisee':
        date_debut_str = request.GET.get('date_debut_custom', '')
        date_fin_str   = request.GET.get('date_fin_custom', '')
        try:
            date_debut = dt_date.fromisoformat(date_debut_str)
            date_fin   = dt_date.fromisoformat(date_fin_str)
            if date_fin < date_debut:
                date_fin = date_debut
        except ValueError:
            date_debut = today
            date_fin   = today
    else:
        date_debut = today.replace(day=1)
        date_fin = today

    # Pour Express : récupérer les dates clôturées
    dates_cloturees = []
    if institut.code == 'express':
        dates_cloturees = list(
            ClotureCaisse.objects.filter(
                institut=institut,
                date__range=[date_debut, date_fin],
                cloture=True
            ).values_list('date', flat=True)
        )

        # Si aucune date clôturée, retourner des données vides
        if not dates_cloturees:
            return JsonResponse({
                'institut': institut.nom,
                'employes': [],
                'prestations': []
            })

    # CA par employé
    employes_stats = []
    employes = Employe.objects.filter(institut=institut, actif=True).order_by('ordre_affichage', 'nom')

    for employe in employes:
        paiements_query = Paiement.objects.filter(
            rendez_vous__employe=employe,
            rendez_vous__statut='valide'
        )

        # Pour Express : filtrer par dates clôturées
        if institut.code == 'express':
            paiements_query = paiements_query.filter(rendez_vous__date__in=dates_cloturees)
        else:
            paiements_query = paiements_query.filter(
                rendez_vous__date__gte=date_debut,
                rendez_vous__date__lte=date_fin
            )

        ca = paiements_query.aggregate(total=Sum('montant'))['total'] or 0

        if ca > 0:  # Ne montrer que les employés avec du CA
            employes_stats.append({
                'nom': employe.nom,
                'ca': int(ca)
            })

    # Top prestations
    prestations_query = RendezVous.objects.filter(
        institut=institut,
        statut='valide',
        prestation__isnull=False  # Exclure les RDV sans prestation
    )

    # Pour Express : filtrer par dates clôturées
    if institut.code == 'express':
        prestations_query = prestations_query.filter(date__in=dates_cloturees)
    else:
        prestations_query = prestations_query.filter(
            date__gte=date_debut,
            date__lte=date_fin
        )

    prestations_stats = prestations_query.values(
        'prestation__nom'
    ).annotate(
        nb_rdv=Count('id'),
        ca_total=Sum('prix_total')
    ).order_by('-ca_total')[:10]  # Top 10 prestations

    prestations_data = [
        {
            'nom': p['prestation__nom'] or 'Sans nom',
            'nb_rdv': p['nb_rdv'],
            'ca': float(p['ca_total']) if p['ca_total'] else 0
        }
        for p in prestations_stats
    ]

    # CA par moyen de paiement
    paiements_query = Paiement.objects.filter(
        rendez_vous__institut=institut,
        rendez_vous__statut='valide'
    ).exclude(mode__in=['carte_cadeau', 'forfait', 'offert'])

    # Pour Express : filtrer par dates clôturées
    if institut.code == 'express':
        paiements_query = paiements_query.filter(rendez_vous__date__in=dates_cloturees)
    else:
        paiements_query = paiements_query.filter(
            rendez_vous__date__gte=date_debut,
            rendez_vous__date__lte=date_fin
        )

    paiements_stats = paiements_query.values('mode').annotate(
        ca_total=Sum('montant')
    ).order_by('-ca_total')

    # Mapper les noms des modes de paiement
    mode_display_map = {
        'especes': 'Espèces',
        'carte': 'Carte',
        'cheque': 'Chèque',
        'om': 'Orange Money',
        'wave': 'Wave',
        'carte_cadeau': 'Carte cadeau',
        'forfait': 'Forfait',
        'offert': 'Offert',
    }

    # Combiner paiements RDV + paiements crédits par mode
    ca_par_mode = {}
    for p in paiements_stats:
        if p['ca_total'] and p['ca_total'] > 0:
            mode = p['mode']
            ca_par_mode[mode] = ca_par_mode.get(mode, 0) + float(p['ca_total'])

    # Ajouter les paiements de crédits pour cet institut
    credits_paiements_qs = PaiementCredit.objects.filter(
        credit__institut=institut
    ).exclude(mode='carte_cadeau')
    if institut.code == 'express':
        credits_paiements_qs = credits_paiements_qs.filter(date__date__in=dates_cloturees)
    else:
        credits_paiements_qs = credits_paiements_qs.filter(
            date__date__gte=date_debut,
            date__date__lte=date_fin
        )
    credits_par_mode = credits_paiements_qs.values('mode').annotate(
        ca_total=Sum('montant')
    )
    for p in credits_par_mode:
        if p['ca_total'] and p['ca_total'] > 0:
            mode = p['mode']
            ca_par_mode[mode] = ca_par_mode.get(mode, 0) + float(p['ca_total'])

    # Ajouter les paiements des cartes cadeaux vendues
    cartes_query = CarteCadeau.objects.filter(
        institut_achat=institut,
        statut__in=['active', 'soldee']
    )
    if institut.code == 'express':
        cartes_query = cartes_query.filter(date_achat__date__in=dates_cloturees)
    else:
        cartes_query = cartes_query.filter(
            date_achat__date__gte=date_debut,
            date_achat__date__lte=date_fin
        )
    for carte in cartes_query:
        mode1 = carte.mode_paiement_achat
        montant1 = carte.montant_paiement_1 if carte.montant_paiement_1 else carte.montant_initial
        ca_par_mode[mode1] = ca_par_mode.get(mode1, 0) + float(montant1)
        if carte.moyen_paiement_2 and carte.montant_paiement_2:
            ca_par_mode[carte.moyen_paiement_2] = ca_par_mode.get(carte.moyen_paiement_2, 0) + float(carte.montant_paiement_2)

    paiements_data = sorted(
        [
            {'mode': mode_display_map.get(mode, mode), 'ca': ca}
            for mode, ca in ca_par_mode.items()
        ],
        key=lambda x: -x['ca']
    )

    return JsonResponse({
        'institut': institut.nom,
        'employes': employes_stats,
        'prestations': prestations_data,
        'paiements': paiements_data
    })


@login_required
@role_required(['patron'])
def export_rdv_excel(request):
    """Export des RDV en Excel."""
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse("Module openpyxl non installé. Installez-le avec: pip install openpyxl", status=500)

    # Paramètres de filtrage
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    institut_code = request.GET.get('institut')

    # Construire le queryset
    rdvs = RendezVous.objects.filter(statut='valide').select_related(
        'client', 'employe', 'institut', 'prestation'
    ).order_by('-date', '-heure_debut')

    if date_debut:
        rdvs = rdvs.filter(date__gte=date_debut)
    if date_fin:
        rdvs = rdvs.filter(date__lte=date_fin)
    if institut_code:
        rdvs = rdvs.filter(institut__code=institut_code)

    # Créer le workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rendez-vous"

    # En-têtes
    headers = ['Date', 'Heure', 'Institut', 'Client', 'Téléphone', 'Employé', 'Prestation', 'Montant (CFA)', 'Mode paiement']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)

    # Données
    row = 2
    for rdv in rdvs:
        modes = ', '.join(dict.fromkeys(p.get_mode_display() for p in rdv.paiements.all()))

        ws.cell(row=row, column=1, value=rdv.date.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=2, value=rdv.heure_debut.strftime('%H:%M'))
        ws.cell(row=row, column=3, value=rdv.institut.nom)
        ws.cell(row=row, column=4, value=rdv.client.get_full_name())
        ws.cell(row=row, column=5, value=rdv.client.telephone)
        ws.cell(row=row, column=6, value=rdv.employe.nom)
        ws.cell(row=row, column=7, value=rdv.prestation.nom if rdv.prestation else '')
        ws.cell(row=row, column=8, value=int(rdv.prix_total))
        ws.cell(row=row, column=9, value=modes)
        row += 1

    # Ajuster la largeur des colonnes
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    # Générer la réponse
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=rdv_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)

    return response
