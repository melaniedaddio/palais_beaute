from django.shortcuts import render, get_object_or_404, redirect
from core.decorators import login_required_json as login_required
from django.http import JsonResponse, HttpResponse
from django.db.models import Q, Max, Count
from django.views.decorators.http import require_POST
from core.decorators import role_required, institut_required
from core.models import (
    Institut, FamillePrestation, Prestation, Option,
    Employe, Presence, TypeAbsence, Absence, Avertissement,
    HoraireEmploye, ModificationPointage,
    TypePrime, Prime, Avance, CalculSalaire,
    CategorieProduit, UniteMesure, Fournisseur, Produit, MouvementStock,
    Inventaire, LigneInventaire,
    CategoriDepense, Depense, DepenseRecurrente, ValidationDepenseRecurrente,
    Paiement, RendezVous,
    VenteProduit, LigneVenteProduit, ReconciliationCaisse,
    Client,
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.utils import timezone
from datetime import date, timedelta
import json


@login_required
@role_required(['patron'])
def catalogue_view(request):
    """
    Vue principale du catalogue des prestations.
    """
    instituts = Institut.objects.all().order_by('nom')

    # Institut sélectionné (par défaut le premier)
    institut_code = request.GET.get('institut', 'palais')
    institut_actif = Institut.objects.filter(code=institut_code).first()

    if not institut_actif:
        institut_actif = instituts.first()

    # Familles avec prestations pour l'institut sélectionné
    familles = FamillePrestation.objects.filter(
        institut=institut_actif
    ).prefetch_related(
        'prestations'
    ).order_by('ordre_affichage', 'nom')

    # Recherche
    recherche = request.GET.get('q', '').strip()
    if recherche:
        familles = familles.filter(
            Q(nom__icontains=recherche) |
            Q(prestations__nom__icontains=recherche)
        ).distinct()

    return render(request, 'gestion/catalogue.html', {
        'instituts': instituts,
        'institut_actif': institut_actif,
        'familles': familles,
        'recherche': recherche,
    })


@login_required
@role_required(['patron'])
@require_POST
def creer_famille(request):
    """
    Créer une nouvelle famille de prestations.
    """
    try:
        data = json.loads(request.body)

        institut = Institut.objects.get(id=data['institut_id'])

        # Vérifier que le nom n'existe pas déjà dans cet institut
        if FamillePrestation.objects.filter(
            nom__iexact=data['nom'],
            institut=institut
        ).exists():
            return JsonResponse({
                'success': False,
                'error': 'Une famille avec ce nom existe déjà dans cet institut'
            })

        # Déterminer l'ordre (dernière position)
        dernier_ordre = FamillePrestation.objects.filter(
            institut=institut
        ).aggregate(Max('ordre_affichage'))['ordre_affichage__max'] or 0

        famille = FamillePrestation.objects.create(
            nom=data['nom'],
            institut=institut,
            couleur=data.get('couleur', '#3498db'),
            ordre_affichage=dernier_ordre + 1
        )

        return JsonResponse({
            'success': True,
            'famille': {
                'id': famille.id,
                'nom': famille.nom,
                'couleur': famille.couleur
            }
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@role_required(['patron'])
@require_POST
def modifier_famille(request, famille_id):
    """
    Modifier une famille existante.
    """
    try:
        data = json.loads(request.body)
        famille = get_object_or_404(FamillePrestation, id=famille_id)

        # Vérifier unicité du nom si changé
        if data.get('nom') and data['nom'].lower() != famille.nom.lower():
            if FamillePrestation.objects.filter(
                nom__iexact=data['nom'],
                institut=famille.institut
            ).exists():
                return JsonResponse({
                    'success': False,
                    'error': 'Une famille avec ce nom existe déjà'
                })

        # Mettre à jour
        if 'nom' in data:
            famille.nom = data['nom']
        if 'couleur' in data:
            famille.couleur = data['couleur']

        famille.save()

        return JsonResponse({
            'success': True,
            'famille': {
                'id': famille.id,
                'nom': famille.nom,
                'couleur': famille.couleur
            }
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@role_required(['patron'])
@require_POST
def supprimer_famille(request, famille_id):
    """
    Supprimer une famille et toutes ses prestations.
    """
    try:
        data = json.loads(request.body)
        famille = get_object_or_404(FamillePrestation, id=famille_id)

        # Vérifier la confirmation
        if data.get('confirmation') != 'SUPPRIMER':
            return JsonResponse({
                'success': False,
                'error': 'Veuillez taper "SUPPRIMER" pour confirmer'
            })

        nom_famille = famille.nom
        nb_prestations = famille.prestations.count()

        # Supprimer (cascade sur les prestations)
        famille.delete()

        return JsonResponse({
            'success': True,
            'message': f'Famille "{nom_famille}" et {nb_prestations} prestation(s) supprimées'
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@role_required(['patron'])
def famille_info_suppression(request, famille_id):
    """
    Récupère les informations avant suppression d'une famille.
    """
    famille = get_object_or_404(FamillePrestation, id=famille_id)
    nb_prestations = famille.prestations.count()

    return JsonResponse({
        'famille': {
            'id': famille.id,
            'nom': famille.nom,
        },
        'nb_prestations': nb_prestations
    })


@login_required
@role_required(['patron'])
@require_POST
def deplacer_famille(request, famille_id, direction):
    """
    Monter ou descendre une famille.
    direction: 'up' ou 'down'
    """
    try:
        famille = get_object_or_404(FamillePrestation, id=famille_id)

        if direction == 'up':
            # Trouver la famille au-dessus
            famille_dessus = FamillePrestation.objects.filter(
                institut=famille.institut,
                ordre_affichage__lt=famille.ordre_affichage
            ).order_by('-ordre_affichage').first()

            if famille_dessus:
                # Échanger les ordres
                famille.ordre_affichage, famille_dessus.ordre_affichage = (
                    famille_dessus.ordre_affichage, famille.ordre_affichage
                )
                famille.save()
                famille_dessus.save()

        elif direction == 'down':
            # Trouver la famille en-dessous
            famille_dessous = FamillePrestation.objects.filter(
                institut=famille.institut,
                ordre_affichage__gt=famille.ordre_affichage
            ).order_by('ordre_affichage').first()

            if famille_dessous:
                famille.ordre_affichage, famille_dessous.ordre_affichage = (
                    famille_dessous.ordre_affichage, famille.ordre_affichage
                )
                famille.save()
                famille_dessous.save()

        return JsonResponse({'success': True})

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


# =============================================================================
# GESTION DES PRESTATIONS
# =============================================================================

@login_required
@role_required(['patron'])
@require_POST
def creer_prestation(request):
    """
    Créer une nouvelle prestation.
    """
    try:
        data = json.loads(request.body)

        famille = FamillePrestation.objects.get(id=data['famille_id'])

        # Vérifier que le nom n'existe pas déjà dans cette famille
        if Prestation.objects.filter(
            nom__iexact=data['nom'],
            famille=famille
        ).exists():
            return JsonResponse({
                'success': False,
                'error': 'Une prestation avec ce nom existe déjà dans cette famille'
            })

        # Vérifier que la durée est fournie pour les prestations normales et forfaits
        type_prestation = data.get('type_prestation', 'normal')
        if type_prestation != 'option' and not data.get('duree_minutes'):
            return JsonResponse({
                'success': False,
                'error': 'La durée est obligatoire pour ce type de prestation'
            })

        # Déterminer l'ordre (dernière position)
        dernier_ordre = Prestation.objects.filter(
            famille=famille
        ).aggregate(Max('ordre_affichage'))['ordre_affichage__max'] or 0

        # Créer la prestation
        prestation = Prestation.objects.create(
            nom=data['nom'],
            famille=famille,
            type_prestation=type_prestation,
            prix=data.get('prix', 0),
            duree_minutes=data.get('duree_minutes'),
            unite=data.get('unite', ''),
            nombre_seances=data.get('nb_seances', 1),
            ordre_affichage=dernier_ordre + 1
        )

        return JsonResponse({
            'success': True,
            'prestation': {
                'id': prestation.id,
                'nom': prestation.nom,
                'type_prestation': prestation.type_prestation,
                'prix': float(prestation.prix),
                'duree_minutes': prestation.duree_minutes,
                'unite': prestation.unite,
                'nb_seances': prestation.nombre_seances,
                'actif': prestation.actif
            }
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@role_required(['patron'])
@require_POST
def modifier_prestation(request, prestation_id):
    """
    Modifier une prestation existante.
    """
    try:
        data = json.loads(request.body)
        prestation = get_object_or_404(Prestation, id=prestation_id)

        # Vérifier unicité du nom si changé
        if data.get('nom') and data['nom'].lower() != prestation.nom.lower():
            if Prestation.objects.filter(
                nom__iexact=data['nom'],
                famille=prestation.famille
            ).exists():
                return JsonResponse({
                    'success': False,
                    'error': 'Une prestation avec ce nom existe déjà dans cette famille'
                })

        # Vérifier que la durée est fournie pour les prestations normales et forfaits
        type_prestation = data.get('type_prestation', prestation.type_prestation)
        if type_prestation != 'option' and 'duree_minutes' in data and not data.get('duree_minutes'):
            return JsonResponse({
                'success': False,
                'error': 'La durée est obligatoire pour ce type de prestation'
            })

        # Mettre à jour les champs
        if 'nom' in data:
            prestation.nom = data['nom']
        if 'type_prestation' in data:
            prestation.type_prestation = data['type_prestation']
        if 'prix' in data:
            prestation.prix = data['prix']
        if 'duree_minutes' in data:
            prestation.duree_minutes = data['duree_minutes']
        if 'unite' in data:
            prestation.unite = data['unite']
        if 'nb_seances' in data:
            prestation.nombre_seances = data['nb_seances']

        prestation.save()

        return JsonResponse({
            'success': True,
            'prestation': {
                'id': prestation.id,
                'nom': prestation.nom,
                'type_prestation': prestation.type_prestation,
                'prix': float(prestation.prix),
                'duree_minutes': prestation.duree_minutes,
                'unite': prestation.unite,
                'nb_seances': prestation.nombre_seances,
                'actif': prestation.actif
            }
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@role_required(['patron'])
@require_POST
def supprimer_prestation(request, prestation_id):
    """
    Supprimer une prestation.
    """
    try:
        data = json.loads(request.body)
        prestation = get_object_or_404(Prestation, id=prestation_id)

        # Vérifier la confirmation
        if data.get('confirmation') != 'SUPPRIMER':
            return JsonResponse({
                'success': False,
                'error': 'Veuillez taper "SUPPRIMER" pour confirmer'
            })

        nom_prestation = prestation.nom

        # Supprimer
        prestation.delete()

        return JsonResponse({
            'success': True,
            'message': f'Prestation "{nom_prestation}" supprimée'
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@role_required(['patron'])
def prestation_info_suppression(request, prestation_id):
    """
    Récupère les informations avant suppression d'une prestation.
    """
    prestation = get_object_or_404(Prestation, id=prestation_id)

    # Vérifier si la prestation est utilisée dans des réservations ou paiements
    # Pour l'instant on retourne juste les infos de base
    return JsonResponse({
        'prestation': {
            'id': prestation.id,
            'nom': prestation.nom,
            'famille': prestation.famille.nom
        }
    })


@login_required
@role_required(['patron'])
def prestation_details(request, prestation_id):
    """
    Récupère toutes les informations d'une prestation pour l'édition.
    """
    prestation = get_object_or_404(Prestation, id=prestation_id)

    # Récupérer les IDs des instituts associés
    instituts_ids = list(prestation.instituts.values_list('id', flat=True))

    return JsonResponse({
        'success': True,
        'prestation': {
            'id': prestation.id,
            'nom': prestation.nom,
            'famille_id': prestation.famille.id,
            'type_prestation': prestation.type_prestation,
            'prix': prestation.prix,
            'duree_minutes': prestation.duree_minutes,
            'unite': prestation.unite or '',
            'nb_seances': prestation.nombre_seances,
            'instituts': instituts_ids,
            'actif': prestation.actif
        }
    })


@login_required
@role_required(['patron'])
@require_POST
def toggle_prestation_actif(request, prestation_id):
    """
    Activer/désactiver une prestation.
    """
    try:
        prestation = get_object_or_404(Prestation, id=prestation_id)
        prestation.actif = not prestation.actif
        prestation.save()

        return JsonResponse({
            'success': True,
            'actif': prestation.actif
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@role_required(['patron'])
@require_POST
def deplacer_prestation(request, prestation_id, direction):
    """
    Monter ou descendre une prestation.
    direction: 'up' ou 'down'
    """
    try:
        prestation = get_object_or_404(Prestation, id=prestation_id)

        if direction == 'up':
            # Trouver la prestation au-dessus
            prestation_dessus = Prestation.objects.filter(
                famille=prestation.famille,
                ordre_affichage__lt=prestation.ordre_affichage
            ).order_by('-ordre_affichage').first()

            if prestation_dessus:
                # Échanger les ordres
                prestation.ordre_affichage, prestation_dessus.ordre_affichage = (
                    prestation_dessus.ordre_affichage, prestation.ordre_affichage
                )
                prestation.save()
                prestation_dessus.save()

        elif direction == 'down':
            # Trouver la prestation en-dessous
            prestation_dessous = Prestation.objects.filter(
                famille=prestation.famille,
                ordre_affichage__gt=prestation.ordre_affichage
            ).order_by('ordre_affichage').first()

            if prestation_dessous:
                prestation.ordre_affichage, prestation_dessous.ordre_affichage = (
                    prestation_dessous.ordre_affichage, prestation.ordre_affichage
                )
                prestation.save()
                prestation_dessous.save()

        return JsonResponse({'success': True})

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


# =============================================================================
# RÉORDONNANCEMENT DRAG & DROP
# =============================================================================

@login_required
@role_required(['patron'])
@require_POST
def reordonner_familles(request, institut_id):
    """
    Réorganiser l'ordre des familles après drag & drop.
    Reçoit une liste d'IDs dans le nouvel ordre.
    """
    try:
        data = json.loads(request.body)
        famille_ids = data.get('famille_ids', [])

        # Vérifier que l'institut existe
        institut = get_object_or_404(Institut, id=institut_id)

        # Mettre à jour l'ordre d'affichage de chaque famille
        for index, famille_id in enumerate(famille_ids):
            FamillePrestation.objects.filter(
                id=famille_id,
                institut=institut
            ).update(ordre_affichage=index + 1)

        return JsonResponse({'success': True})

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@role_required(['patron'])
@require_POST
def reordonner_prestations(request, famille_id):
    """
    Réorganiser l'ordre des prestations dans une famille après drag & drop.
    Reçoit une liste d'IDs dans le nouvel ordre.
    """
    try:
        data = json.loads(request.body)
        prestation_ids = data.get('prestation_ids', [])

        # Vérifier que la famille existe
        famille = get_object_or_404(FamillePrestation, id=famille_id)

        # Mettre à jour l'ordre d'affichage de chaque prestation
        for index, prestation_id in enumerate(prestation_ids):
            Prestation.objects.filter(
                id=prestation_id,
                famille=famille
            ).update(ordre_affichage=index + 1)

        return JsonResponse({'success': True})

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


# =============================================================================
# EXPORT EXCEL
# =============================================================================

@login_required
@role_required(['patron'])
def export_catalogue_excel(request):
    """
    Exporter le catalogue des prestations en Excel.
    """
    try:
        # Institut sélectionné
        institut_code = request.GET.get('institut', 'palais')
        institut = Institut.objects.filter(code=institut_code).first()

        if not institut:
            institut = Institut.objects.first()

        # Créer le workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Catalogue {institut.nom}"

        # Styles
        header_fill = PatternFill(start_color="E8B4B8", end_color="E8B4B8", fill_type="solid")
        famille_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        header_font = Font(bold=True, size=12, color="FFFFFF")
        famille_font = Font(bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # En-tête du document
        ws['A1'] = f"CATALOGUE DES PRESTATIONS - {institut.nom.upper()}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:G1')

        ws['A2'] = f"Généré le {timezone.now().strftime('%d/%m/%Y à %H:%M')}"
        ws['A2'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A2:G2')

        # En-têtes des colonnes
        row = 4
        headers = ['Famille', 'Prestation', 'Type', 'Prix (CFA)', 'Durée', 'Unité', 'Statut']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Données
        row = 5
        familles = FamillePrestation.objects.filter(
            institut=institut,
            actif=True
        ).prefetch_related('prestations').order_by('ordre_affichage', 'nom')

        for famille in familles:
            prestations = famille.prestations.order_by('ordre_affichage', 'nom')

            if not prestations.exists():
                # Famille sans prestations
                ws.cell(row=row, column=1, value=famille.nom).font = famille_font
                ws.cell(row=row, column=1).fill = famille_fill
                ws.cell(row=row, column=2, value="(Aucune prestation)")
                for col in range(1, 8):
                    ws.cell(row=row, column=col).border = border
                row += 1
            else:
                # Prestations de la famille
                for idx, prestation in enumerate(prestations):
                    # Afficher le nom de la famille uniquement sur la première ligne
                    if idx == 0:
                        cell = ws.cell(row=row, column=1, value=famille.nom)
                        cell.font = famille_font
                        cell.fill = famille_fill
                    else:
                        cell = ws.cell(row=row, column=1, value="")
                        cell.fill = famille_fill

                    # Type de prestation
                    if prestation.type_prestation == 'forfait':
                        type_display = f"Forfait ({prestation.nombre_seances} séances)"
                    elif prestation.type_prestation == 'option':
                        type_display = "Option"
                    else:
                        type_display = "Normal"

                    # Données de la prestation
                    ws.cell(row=row, column=2, value=prestation.nom)
                    ws.cell(row=row, column=3, value=type_display)
                    ws.cell(row=row, column=4, value=prestation.prix)
                    ws.cell(row=row, column=5, value=prestation.get_duree_display())
                    ws.cell(row=row, column=6, value=prestation.unite or "-")
                    ws.cell(row=row, column=7, value="Actif" if prestation.actif else "Inactif")

                    # Bordures
                    for col in range(1, 8):
                        ws.cell(row=row, column=col).border = border

                    row += 1

        # Ajuster les largeurs de colonnes
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 12

        # Générer la réponse
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'catalogue_{institut.code}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)

        return response

    except Exception as e:
        return HttpResponse(f"Erreur lors de l'export: {str(e)}", status=500)

# === Gestion des Options ===

@login_required
@institut_required
def api_options_liste(request, institut_code):
    """API : Liste des options d'un institut"""
    institut = get_object_or_404(Institut, code=institut_code)
    
    options = Option.objects.filter(institut=institut).order_by('nom')
    
    options_data = []
    for option in options:
        options_data.append({
            'id': option.id,
            'nom': option.nom,
            'prix': option.prix,
            'a_quantite': option.a_quantite,
            'unite': option.unite or '',
            'actif': option.actif,
        })
    
    return JsonResponse({
        'success': True,
        'options': options_data
    })


@login_required
@institut_required
@require_POST
def api_option_creer(request, institut_code):
    """API : Créer une nouvelle option"""
    institut = get_object_or_404(Institut, code=institut_code)
    
    try:
        nom = request.POST.get('nom', '').strip()
        prix = int(request.POST.get('prix', 0))
        a_quantite = request.POST.get('a_quantite') == 'true'
        unite = request.POST.get('unite', '').strip()
        
        if not nom:
            return JsonResponse({
                'success': False,
                'message': 'Le nom est requis'
            }, status=400)
        
        if prix < 0:
            return JsonResponse({
                'success': False,
                'message': 'Le prix ne peut pas être négatif'
            }, status=400)
        
        option = Option.objects.create(
            nom=nom,
            institut=institut,
            prix=prix,
            a_quantite=a_quantite,
            unite=unite if unite else None,
            actif=True
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Option "{nom}" créée avec succès',
            'option': {
                'id': option.id,
                'nom': option.nom,
                'prix': option.prix,
                'a_quantite': option.a_quantite,
                'unite': option.unite or '',
                'actif': option.actif,
            }
        })
        
    except ValueError:
        return JsonResponse({
            'success': False,
            'message': 'Prix invalide'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@login_required
@institut_required
@require_POST
def api_option_modifier(request, institut_code, option_id):
    """API : Modifier une option"""
    institut = get_object_or_404(Institut, code=institut_code)
    option = get_object_or_404(Option, id=option_id, institut=institut)
    
    try:
        option.nom = request.POST.get('nom', option.nom).strip()
        option.prix = int(request.POST.get('prix', option.prix))
        option.a_quantite = request.POST.get('a_quantite') == 'true'
        unite = request.POST.get('unite', '').strip()
        option.unite = unite if unite else None
        option.actif = request.POST.get('actif', 'true') == 'true'
        
        option.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Option "{option.nom}" modifiée avec succès',
            'option': {
                'id': option.id,
                'nom': option.nom,
                'prix': option.prix,
                'a_quantite': option.a_quantite,
                'unite': option.unite or '',
                'actif': option.actif,
            }
        })
        
    except ValueError:
        return JsonResponse({
            'success': False,
            'message': 'Prix invalide'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@login_required
@institut_required
@require_POST
def api_option_supprimer(request, institut_code, option_id):
    """API : Supprimer une option"""
    institut = get_object_or_404(Institut, code=institut_code)
    option = get_object_or_404(Option, id=option_id, institut=institut)
    
    try:
        nom = option.nom
        option.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Option "{nom}" supprimée avec succès'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@login_required
@institut_required
def api_option_details(request, institut_code, option_id):
    """API : Détails d'une option"""
    institut = get_object_or_404(Institut, code=institut_code)
    option = get_object_or_404(Option, id=option_id, institut=institut)
    
    return JsonResponse({
        'success': True,
        'option': {
            'id': option.id,
            'nom': option.nom,
            'prix': option.prix,
            'a_quantite': option.a_quantite,
            'unite': option.unite or '',
            'actif': option.actif,
        }
    })


# ============================================================
# VUES PRÉSENCES
# ============================================================

@login_required
@role_required(['patron', 'manager'])
def presences_pointage(request):
    """Page de pointage du jour."""
    utilisateur = request.user.utilisateur
    today = timezone.localtime(timezone.now()).date()

    if utilisateur.is_patron():
        instituts = list(Institut.objects.all().order_by('nom'))
        institut_code = request.GET.get('institut', instituts[0].code if instituts else 'palais')
    else:
        instituts = [utilisateur.institut]
        institut_code = utilisateur.institut.code

    if institut_code == 'autres':
        institut_actif = None
        employes = Employe.objects.filter(institut__isnull=True, actif=True).order_by('ordre_affichage', 'nom')
    else:
        institut_actif = Institut.objects.filter(code=institut_code).first()
        employes = Employe.objects.filter(institut=institut_actif, actif=True).order_by('ordre_affichage', 'nom')

    presences = {p.employe_id: p for p in Presence.objects.filter(employe__in=employes, date=today)}

    # Horaires théoriques par employé pour affichage et vérification JS
    horaires = {}
    for emp in employes:
        hd, hf = HoraireEmploye.get_horaire_pour_date(emp, today)
        horaires[emp.id] = {'debut': hd.strftime('%H:%M'), 'fin': hf.strftime('%H:%M')}

    return render(request, 'gestion/presences/pointage.html', {
        'instituts': instituts,
        'institut_actif': institut_actif,
        'institut_code': institut_code,
        'employes': employes,
        'presences': presences,
        'horaires': horaires,
        'today': today,
        'is_patron': utilisateur.is_patron(),
    })


@login_required
@role_required(['patron', 'manager'])
def presences_historique(request):
    """Historique unifié des présences (remplace absences + retards + ancien historique)."""
    utilisateur = request.user.utilisateur

    # Période : semaine / mois / custom
    periode = request.GET.get('periode', 'mois')  # 'semaine', 'mois', 'custom'
    today = date.today()

    if periode == 'semaine':
        d_debut = today - timedelta(days=today.weekday())
        d_fin = d_debut + timedelta(days=6)
    elif periode == 'custom':
        try:
            d_debut = date.fromisoformat(request.GET.get('date_debut', ''))
            d_fin = date.fromisoformat(request.GET.get('date_fin', ''))
        except ValueError:
            d_debut = today.replace(day=1)
            d_fin = today
    else:  # mois
        mois_str = request.GET.get('mois')
        if mois_str:
            try:
                mois = date.fromisoformat(f"{mois_str}-01")
            except ValueError:
                mois = today.replace(day=1)
        else:
            mois = today.replace(day=1)
        d_debut = mois
        d_fin = (mois.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)

    if utilisateur.is_patron():
        instituts = list(Institut.objects.all().order_by('nom'))
        institut_code = request.GET.get('institut', 'tous')
    else:
        instituts = [utilisateur.institut]
        institut_code = utilisateur.institut.code

    employe_id_filtre = request.GET.get('employe_id', '')
    statut_filtre = request.GET.get('statut', 'tous')  # 'tous', 'retard', 'absent', 'absent_nj', 'valide'

    presences_qs = Presence.objects.filter(
        date__range=[d_debut, d_fin]
    ).select_related('employe', 'employe__categorie', 'employe__institut', 'valide_par')

    if institut_code == 'autres':
        presences_qs = presences_qs.filter(employe__institut__isnull=True)
    elif institut_code != 'tous':
        presences_qs = presences_qs.filter(employe__institut__code=institut_code)

    if utilisateur.is_manager():
        presences_qs = presences_qs.filter(employe__institut=utilisateur.institut)

    if employe_id_filtre:
        presences_qs = presences_qs.filter(employe_id=employe_id_filtre)

    if statut_filtre == 'retard':
        presences_qs = presences_qs.filter(statut_arrivee='retard')
    elif statut_filtre == 'absent':
        presences_qs = presences_qs.filter(statut_arrivee='absent')
    elif statut_filtre == 'absent_nj':
        presences_qs = presences_qs.filter(statut_journee='absent_non_justifie')
    elif statut_filtre == 'valide':
        presences_qs = presences_qs.filter(est_valide=True)

    # Stats résumé
    nb_total = presences_qs.count()
    nb_retards = presences_qs.filter(statut_arrivee='retard').count()
    nb_absents = presences_qs.filter(statut_arrivee='absent').count()
    nb_valides = presences_qs.filter(est_valide=True).count()

    # Liste des employés pour le filtre
    employe_filter_q = {}
    if institut_code == 'autres':
        employe_filter_q['institut__isnull'] = True
    elif institut_code != 'tous':
        employe_filter_q['institut__code'] = institut_code
    if utilisateur.is_manager():
        employe_filter_q['institut'] = utilisateur.institut
    employes_filtre = Employe.objects.filter(actif=True, **employe_filter_q).order_by('nom')

    return render(request, 'gestion/presences/historique.html', {
        'instituts': instituts,
        'institut_code': institut_code,
        'periode': periode,
        'd_debut': d_debut,
        'd_fin': d_fin,
        'presences': presences_qs.order_by('-date', 'employe__nom'),
        'nb_total': nb_total,
        'nb_retards': nb_retards,
        'nb_absents': nb_absents,
        'nb_valides': nb_valides,
        'employes_filtre': employes_filtre,
        'employe_id_filtre': employe_id_filtre,
        'statut_filtre': statut_filtre,
        'is_patron': utilisateur.is_patron(),
    })


@login_required
@role_required(['patron', 'manager'])
def absences_liste(request):
    """Liste et gestion des absences."""
    utilisateur = request.user.utilisateur

    mois_str = request.GET.get('mois')
    if mois_str:
        try:
            mois = date.fromisoformat(f"{mois_str}-01")
        except ValueError:
            mois = date.today().replace(day=1)
    else:
        mois = date.today().replace(day=1)

    fin_mois = (mois.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)

    if utilisateur.is_patron():
        instituts = list(Institut.objects.all().order_by('nom'))
        institut_code = request.GET.get('institut', 'tous')
        absences = Absence.objects.filter(
            date_debut__lte=fin_mois, date_fin__gte=mois
        ).select_related('employe', 'employe__institut', 'type_absence')
    else:
        instituts = [utilisateur.institut]
        institut_code = utilisateur.institut.code
        absences = Absence.objects.filter(
            employe__institut=utilisateur.institut,
            date_debut__lte=fin_mois, date_fin__gte=mois
        ).select_related('employe', 'type_absence')

    if institut_code not in ('tous', '') and utilisateur.is_patron():
        if institut_code == 'autres':
            absences = absences.filter(employe__institut__isnull=True)
        else:
            absences = absences.filter(employe__institut__code=institut_code)

    types_absence = TypeAbsence.objects.filter(actif=True)

    if utilisateur.is_patron():
        employes = Employe.objects.filter(actif=True).order_by('nom')
    else:
        employes = Employe.objects.filter(institut=utilisateur.institut, actif=True).order_by('nom')

    return render(request, 'gestion/presences/absences.html', {
        'absences': absences.order_by('-date_debut'),
        'types_absence': types_absence,
        'employes': employes,
        'instituts': instituts,
        'institut_code': institut_code,
        'mois': mois,
        'is_patron': utilisateur.is_patron(),
    })


@login_required
@role_required(['patron', 'manager'])
def retards_suivi(request):
    """Suivi des retards avec possibilité de créer des avertissements."""
    utilisateur = request.user.utilisateur

    mois_str = request.GET.get('mois')
    if mois_str:
        try:
            mois = date.fromisoformat(f"{mois_str}-01")
        except ValueError:
            mois = date.today().replace(day=1)
    else:
        mois = date.today().replace(day=1)

    fin_mois = (mois.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)

    if utilisateur.is_patron():
        instituts = list(Institut.objects.all().order_by('nom'))
        institut_code = request.GET.get('institut', 'tous')
        employes_qs = Employe.objects.filter(actif=True)
    else:
        instituts = [utilisateur.institut]
        institut_code = utilisateur.institut.code
        employes_qs = Employe.objects.filter(institut=utilisateur.institut, actif=True)

    if institut_code not in ('tous', '') and utilisateur.is_patron():
        if institut_code == 'autres':
            employes_qs = employes_qs.filter(institut__isnull=True)
        else:
            employes_qs = employes_qs.filter(institut__code=institut_code)

    # Retards du mois par employé
    retards_par_employe = []
    for emp in employes_qs.order_by('nom'):
        nb_retards = Presence.objects.filter(
            employe=emp, date__range=[mois, fin_mois], statut_arrivee='retard'
        ).count()
        if nb_retards > 0:
            avert_existant = Avertissement.objects.filter(
                employe=emp, mois_concerne=mois,
                type_avertissement='retard'
            ).first()
            retards_par_employe.append({
                'employe': emp,
                'nb_retards': nb_retards,
                'avertissement': avert_existant,
            })

    avertissements_recents = Avertissement.objects.filter(
        mois_concerne__gte=mois - timedelta(days=90)
    ).select_related('employe').order_by('-date_creation')[:20]

    return render(request, 'gestion/presences/retards.html', {
        'instituts': instituts,
        'institut_code': institut_code,
        'mois': mois,
        'retards_par_employe': retards_par_employe,
        'avertissements_recents': avertissements_recents,
        'is_patron': utilisateur.is_patron(),
    })


@login_required
@role_required(['patron', 'manager'])
@require_POST
def api_pointer(request):
    """API pour enregistrer une présence.

    Types supportés : 'arrivee', 'depart', 'depart_midi', 'retour_midi'
    - Arrivée : présent si <= heure_debut+15min (depuis HoraireEmploye), retard sinon
    - Départ  : présent si >= heure_fin, anticipe sinon
    Si statut='absent' est fourni explicitement, on marque absent.
    """
    from datetime import time as dtime, timedelta as td

    TOLERANCE_RETARD_MINUTES = 15

    try:
        data = json.loads(request.body)
        employe = get_object_or_404(Employe, id=data['employe_id'])

        today_abidjan = timezone.localtime(timezone.now()).date()

        # Vérifier si une absence longue couvre la date
        absence_active = Absence.objects.filter(
            employe=employe, date_debut__lte=today_abidjan, date_fin__gte=today_abidjan
        ).first()

        try:
            presence = Presence.objects.get(employe=employe, date=today_abidjan)
        except Presence.DoesNotExist:
            presence = Presence(
                employe=employe,
                date=today_abidjan,
                saisi_par=request.user.utilisateur,
                statut_arrivee='present',
            )

        type_pointage = data.get('type')
        statut_force  = data.get('statut')

        now_local  = timezone.localtime(timezone.now())
        heure_now  = now_local.time()
        heure_str  = heure_now.strftime('%H:%M')

        # Lire l'horaire théorique de l'employé
        heure_debut_theo, heure_fin_theo = HoraireEmploye.get_horaire_pour_date(employe, today_abidjan)

        # Seuil retard = heure_debut + TOLERANCE_RETARD_MINUTES
        from datetime import datetime as dt_cls, date as date_cls
        seuil_retard_dt = dt_cls.combine(date_cls.today(), heure_debut_theo) + td(minutes=TOLERANCE_RETARD_MINUTES)
        seuil_retard = seuil_retard_dt.time()

        if type_pointage == 'arrivee':
            if absence_active:
                presence.statut_arrivee = 'absent'
                presence.heure_arrivee = None
                presence.statut_journee = 'conge_long'
                statut_result = 'absent'
                heure_str = None
            elif statut_force == 'absent':
                presence.statut_arrivee = 'absent'
                presence.heure_arrivee  = None
                presence.statut_journee = 'absent_non_justifie'
                presence.minutes_retard = 0
                # Auto-valider : c'est le patron/manager qui marque manuellement
                presence.est_valide = True
                presence.date_validation = timezone.now()
                statut_result = 'absent'
                heure_str = None
            else:
                is_retard = heure_now > seuil_retard
                statut_result = 'retard' if is_retard else 'present'
                presence.statut_arrivee = statut_result
                presence.heure_arrivee  = heure_now
                if is_retard:
                    debut_dt = dt_cls.combine(date_cls.today(), heure_debut_theo)
                    now_dt = dt_cls.combine(date_cls.today(), heure_now)
                    presence.minutes_retard = int((now_dt - debut_dt).total_seconds() / 60)
                else:
                    presence.minutes_retard = 0

        elif type_pointage == 'depart':
            if statut_force == 'absent':
                presence.statut_depart = 'absent'
                presence.heure_depart  = None
                statut_result = 'absent'
                heure_str = None
            else:
                statut_result = 'anticipe' if heure_now < heure_fin_theo else 'present'
                presence.statut_depart = statut_result
                presence.heure_depart  = heure_now

        elif type_pointage == 'depart_midi':
            presence.heure_depart_midi = heure_now
            statut_result = 'depart_midi'

        elif type_pointage == 'retour_midi':
            presence.heure_retour_midi = heure_now
            statut_result = 'retour_midi'

        else:
            return JsonResponse({'success': False, 'error': 'type invalide'}, status=400)

        presence.save()

        return JsonResponse({
            'success': True,
            'statut': statut_result,
            'heure':  heure_str,
            'heure_debut_theo': heure_debut_theo.strftime('%H:%M'),
            'heure_fin_theo': heure_fin_theo.strftime('%H:%M'),
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
@role_required(['patron', 'manager'])
@require_POST
def api_absence_creer(request):
    """API pour créer une absence."""
    try:
        employe_id = request.POST.get('employe_id')
        type_id = request.POST.get('type_absence_id')
        date_debut_str = request.POST.get('date_debut')
        date_fin_str = request.POST.get('date_fin')

        if not all([employe_id, type_id, date_debut_str, date_fin_str]):
            return JsonResponse({'success': False, 'error': 'Champs obligatoires manquants'}, status=400)

        employe = get_object_or_404(Employe, id=employe_id)
        type_absence = get_object_or_404(TypeAbsence, id=type_id)
        d_debut = date.fromisoformat(date_debut_str)
        d_fin = date.fromisoformat(date_fin_str)

        if d_fin < d_debut:
            return JsonResponse({'success': False, 'error': 'La date de fin doit être après la date de début'}, status=400)

        absence = Absence.objects.create(
            employe=employe,
            type_absence=type_absence,
            date_debut=d_debut,
            date_fin=d_fin,
            justificatif_recu=request.POST.get('justificatif_recu') == 'on',
            commentaire=request.POST.get('commentaire', '').strip(),
            cree_par=request.user.utilisateur,
        )

        return JsonResponse({
            'success': True,
            'message': f'Absence de {employe.get_full_name()} enregistrée ({absence.nombre_jours()} jour(s))',
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
@role_required(['patron', 'manager'])
@require_POST
def api_absence_supprimer(request, absence_id):
    """API pour supprimer une absence."""
    try:
        absence = get_object_or_404(Absence, id=absence_id)
        nom = f"{absence.employe.get_full_name()} ({absence.type_absence.nom})"
        absence.delete()
        return JsonResponse({'success': True, 'message': f'Absence de {nom} supprimée'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
@role_required(['patron', 'manager'])
@require_POST
def api_avertissement_creer(request):
    """API pour créer un avertissement."""
    try:
        employe_id = request.POST.get('employe_id')
        mois_str = request.POST.get('mois_concerne')
        type_avert = request.POST.get('type_avertissement', 'retard')

        employe = get_object_or_404(Employe, id=employe_id)
        mois_concerne = date.fromisoformat(f"{mois_str}-01")

        mise_a_pied = request.POST.get('mise_a_pied') == 'on'
        jours_map = int(request.POST.get('jours_mise_a_pied', 0) or 0)

        avert = Avertissement.objects.create(
            employe=employe,
            type_avertissement=type_avert,
            mois_concerne=mois_concerne,
            nombre_retards=int(request.POST.get('nombre_retards', 0) or 0),
            commentaire=request.POST.get('commentaire', '').strip(),
            mise_a_pied=mise_a_pied,
            jours_mise_a_pied=jours_map if mise_a_pied else 0,
            cree_par=request.user.utilisateur,
        )

        return JsonResponse({
            'success': True,
            'message': f'Avertissement créé pour {employe.get_full_name()}',
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


# ============================================================
# VUES PRÉSENCES — DASHBOARD / VALIDATION / HORAIRES / EXPORT
# ============================================================

@login_required
@role_required(['patron'])
def presences_dashboard(request):
    """Tableau de bord journalier des présences (patron seulement)."""
    utilisateur = request.user.utilisateur
    instituts = list(Institut.objects.all().order_by('nom'))

    date_str = request.GET.get('date')
    if date_str:
        try:
            jour = date.fromisoformat(date_str)
        except ValueError:
            jour = date.today()
    else:
        jour = date.today()

    institut_code = request.GET.get('institut', instituts[0].code if instituts else 'palais')
    if institut_code == 'autres':
        institut_actif = None
        employes = Employe.objects.filter(institut__isnull=True, actif=True).order_by('ordre_affichage', 'nom')
    else:
        institut_actif = Institut.objects.filter(code=institut_code).first()
        employes = Employe.objects.filter(institut=institut_actif, actif=True).order_by('ordre_affichage', 'nom')

    presences_map = {p.employe_id: p for p in Presence.objects.filter(employe__in=employes, date=jour).select_related('valide_par')}

    donnees = []
    nb_presents = nb_absents = nb_retards = nb_valides = 0
    for emp in employes:
        p = presences_map.get(emp.id)
        absence = Absence.objects.filter(employe=emp, date_debut__lte=jour, date_fin__gte=jour).first()
        heure_debut_theo, heure_fin_theo = HoraireEmploye.get_horaire_pour_date(emp, jour)

        if p:
            if p.est_valide:
                nb_valides += 1
            if p.statut_arrivee == 'absent' or p.statut_journee == 'conge_long':
                nb_absents += 1
            elif p.statut_arrivee in ('present', 'retard'):
                nb_presents += 1
                if p.minutes_retard > 0:
                    nb_retards += 1
        elif absence:
            nb_absents += 1

        donnees.append({
            'employe': emp,
            'presence': p,
            'absence_longue': absence,
            'heure_debut_theo': heure_debut_theo.strftime('%H:%M'),
            'heure_fin_theo': heure_fin_theo.strftime('%H:%M'),
        })

    return render(request, 'gestion/presences/dashboard.html', {
        'donnees': donnees,
        'instituts': instituts,
        'institut_actif': institut_actif,
        'institut_code': institut_code,
        'jour': jour,
        'nb_presents': nb_presents,
        'nb_absents': nb_absents,
        'nb_retards': nb_retards,
        'nb_valides': nb_valides,
        'total_employes': len(donnees),
        'is_patron': True,
    })


@login_required
@role_required(['patron'])
@require_POST
def api_valider_journee(request):
    """Valider (qualifier) une journée de présence — patron seulement."""
    try:
        data = json.loads(request.body)
        presence_id = data.get('presence_id')
        statut_journee = data.get('statut_journee', 'present')
        notes = data.get('notes', '')

        presence = get_object_or_404(Presence, id=presence_id)
        utilisateur = request.user.utilisateur

        presence.statut_journee = statut_journee
        presence.est_valide = True
        presence.date_validation = timezone.now()
        # Enregistrer les notes (+ qui a validé si commentaire existait)
        commentaire_parts = []
        if notes:
            commentaire_parts.append(notes)
        commentaire_parts.append(f"Validé par {utilisateur.user.get_full_name() or utilisateur.user.username}")
        presence.commentaire = ' | '.join(commentaire_parts)
        presence.save()

        return JsonResponse({
            'success': True,
            'message': f'Journée de {presence.employe.get_full_name()} validée',
            'statut_journee': statut_journee,
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
@role_required(['patron'])
@require_POST
def api_modifier_pointage(request):
    """Modifier une heure de pointage avec audit trail — patron seulement."""
    CHAMPS_AUTORISES = ('heure_arrivee', 'heure_depart', 'heure_depart_midi', 'heure_retour_midi')
    try:
        data = json.loads(request.body)
        presence_id = data.get('presence_id')
        champ = data.get('champ')
        valeur = data.get('valeur', '').strip()
        motif = data.get('motif', '').strip()

        if champ not in CHAMPS_AUTORISES:
            return JsonResponse({'success': False, 'error': 'Champ non autorisé'}, status=400)

        presence = get_object_or_404(Presence, id=presence_id)
        ancienne_valeur = str(getattr(presence, champ) or '')

        if valeur:
            from datetime import time as dtime
            h, m = map(int, valeur.split(':'))
            setattr(presence, champ, dtime(h, m))
        else:
            setattr(presence, champ, None)
        presence.save()

        ModificationPointage.objects.create(
            presence=presence,
            modifie_par=request.user.utilisateur,
            champ_modifie=champ,
            ancienne_valeur=ancienne_valeur,
            nouvelle_valeur=valeur,
            motif=motif,
        )

        return JsonResponse({'success': True, 'message': 'Pointage modifié'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
@role_required(['patron'])
def export_presences(request):
    """Export CSV des présences sur une période — patron seulement."""
    import csv
    from django.http import HttpResponse

    date_debut_str = request.GET.get('date_debut')
    date_fin_str = request.GET.get('date_fin')
    institut_code = request.GET.get('institut', 'tous')

    try:
        d_debut = date.fromisoformat(date_debut_str) if date_debut_str else date.today().replace(day=1)
        d_fin = date.fromisoformat(date_fin_str) if date_fin_str else date.today()
    except ValueError:
        d_debut = date.today().replace(day=1)
        d_fin = date.today()

    qs = Presence.objects.filter(date__range=[d_debut, d_fin]).select_related(
        'employe', 'employe__institut', 'valide_par'
    ).order_by('date', 'employe__nom')

    if institut_code == 'autres':
        qs = qs.filter(employe__institut__isnull=True)
    elif institut_code != 'tous':
        qs = qs.filter(employe__institut__code=institut_code)

    response = HttpResponse(content_type='text/csv; charset=utf-8')
    filename = f"presences_{d_debut.strftime('%Y%m%d')}_{d_fin.strftime('%Y%m%d')}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.write('\ufeff')  # BOM UTF-8 pour Excel

    writer = csv.writer(response, delimiter=';')
    writer.writerow([
        'Date', 'Employé', 'Institut',
        'Heure arrivée', 'Statut arrivée', 'Retard (min)',
        'Départ midi', 'Retour midi',
        'Heure départ', 'Statut départ',
        'Durée travaillée', 'Statut journée', 'Validé', 'Validé par',
    ])

    for p in qs:
        writer.writerow([
            p.date.strftime('%d/%m/%Y'),
            p.employe.get_full_name(),
            p.employe.institut.nom if p.employe.institut else 'Autres',
            p.heure_arrivee.strftime('%H:%M') if p.heure_arrivee else '',
            p.get_statut_arrivee_display(),
            p.minutes_retard if p.minutes_retard else '',
            p.heure_depart_midi.strftime('%H:%M') if p.heure_depart_midi else '',
            p.heure_retour_midi.strftime('%H:%M') if p.heure_retour_midi else '',
            p.heure_depart.strftime('%H:%M') if p.heure_depart else '',
            p.get_statut_depart_display() if p.statut_depart else '',
            p.duree_travaillee or '',
            p.get_statut_journee_display(),
            'Oui' if p.est_valide else 'Non',
            p.valide_par.get_full_name() if p.valide_par else '',
        ])

    return response


@login_required
@role_required(['patron'])
def horaires_config(request):
    """Configuration des horaires théoriques par employé — patron seulement."""
    instituts = list(Institut.objects.all().order_by('nom'))
    institut_code = request.GET.get('institut', 'tous')

    employes_qs = Employe.objects.filter(actif=True).order_by('nom')
    if institut_code == 'autres':
        employes_qs = employes_qs.filter(institut__isnull=True)
    elif institut_code != 'tous':
        employes_qs = employes_qs.filter(institut__code=institut_code)

    # Pour chaque employé, son horaire actif
    donnees = []
    for emp in employes_qs:
        horaires = HoraireEmploye.objects.filter(employe=emp).order_by('-date_debut')[:5]
        hd, hf = HoraireEmploye.get_horaire_pour_date(emp, date.today())
        donnees.append({
            'employe': emp,
            'horaires': horaires,
            'horaire_actif_debut': hd.strftime('%H:%M'),
            'horaire_actif_fin': hf.strftime('%H:%M'),
        })

    if request.method == 'POST':
        employe_id = request.POST.get('employe_id')
        heure_debut_str = request.POST.get('heure_debut', '08:00')
        heure_fin_str = request.POST.get('heure_fin', '17:00')
        date_debut_str = request.POST.get('date_debut')
        date_fin_str = request.POST.get('date_fin')

        from datetime import time as dtime
        try:
            emp = get_object_or_404(Employe, id=employe_id)
            hd_h, hd_m = map(int, heure_debut_str.split(':'))
            hf_h, hf_m = map(int, heure_fin_str.split(':'))
            d_debut = date.fromisoformat(date_debut_str)
            d_fin = date.fromisoformat(date_fin_str) if date_fin_str else None

            HoraireEmploye.objects.create(
                employe=emp,
                heure_debut=dtime(hd_h, hd_m),
                heure_fin=dtime(hf_h, hf_m),
                date_debut=d_debut,
                date_fin=d_fin,
            )
            from django.contrib import messages
            messages.success(request, f'Horaire créé pour {emp.get_full_name()}')
        except Exception as e:
            from django.contrib import messages
            messages.error(request, f'Erreur : {e}')

        from django.shortcuts import redirect
        return redirect(f'{request.path}?institut={institut_code}')

    return render(request, 'gestion/presences/horaires_config.html', {
        'donnees': donnees,
        'instituts': instituts,
        'institut_code': institut_code,
        'employes': employes_qs,
        'today': date.today().isoformat(),
    })


@login_required
@role_required(['patron'])
@require_POST
def api_horaire_supprimer(request, horaire_id):
    """Supprimer un horaire employé — patron seulement."""
    try:
        horaire = get_object_or_404(HoraireEmploye, id=horaire_id)
        nom = horaire.employe.get_full_name()
        horaire.delete()
        return JsonResponse({'success': True, 'message': f'Horaire supprimé pour {nom}'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
@role_required(['patron'])
def api_horaires_employe(request, employe_id):
    """Retourne les horaires d'un employé sous forme JSON."""
    employe = get_object_or_404(Employe, id=employe_id)
    horaires = HoraireEmploye.objects.filter(employe=employe).order_by('-date_debut')
    heure_actif_debut, heure_actif_fin = HoraireEmploye.get_horaire_pour_date(employe, date.today())
    return JsonResponse({
        'employe': employe.get_full_name(),
        'horaire_actif': f"{heure_actif_debut.strftime('%H:%M')} – {heure_actif_fin.strftime('%H:%M')}",
        'horaires': [
            {
                'id': h.id,
                'heure_debut': h.heure_debut.strftime('%H:%M'),
                'heure_fin': h.heure_fin.strftime('%H:%M'),
                'date_debut': h.date_debut.strftime('%d/%m/%Y'),
                'date_debut_iso': h.date_debut.isoformat(),
                'date_fin': h.date_fin.strftime('%d/%m/%Y') if h.date_fin else None,
                'date_fin_iso': h.date_fin.isoformat() if h.date_fin else None,
                'actif': not h.date_fin or h.date_fin >= date.today(),
            }
            for h in horaires
        ],
    })


@login_required
@role_required(['patron'])
@require_POST
def api_horaire_creer(request):
    """Créer un horaire pour un employé — retourne JSON."""
    from datetime import time as dtime
    try:
        data = json.loads(request.body)
        employe = get_object_or_404(Employe, id=data['employe_id'])
        hd_h, hd_m = map(int, data['heure_debut'].split(':'))
        hf_h, hf_m = map(int, data['heure_fin'].split(':'))
        d_debut = date.fromisoformat(data['date_debut'])
        d_fin = date.fromisoformat(data['date_fin']) if data.get('date_fin') else None

        horaire = HoraireEmploye.objects.create(
            employe=employe,
            heure_debut=dtime(hd_h, hd_m),
            heure_fin=dtime(hf_h, hf_m),
            date_debut=d_debut,
            date_fin=d_fin,
        )
        return JsonResponse({
            'success': True,
            'message': f'Horaire créé pour {employe.get_full_name()}',
            'horaire': {
                'id': horaire.id,
                'heure_debut': horaire.heure_debut.strftime('%H:%M'),
                'heure_fin': horaire.heure_fin.strftime('%H:%M'),
                'date_debut': horaire.date_debut.strftime('%d/%m/%Y'),
                'date_debut_iso': horaire.date_debut.isoformat(),
                'date_fin': horaire.date_fin.strftime('%d/%m/%Y') if horaire.date_fin else None,
                'date_fin_iso': horaire.date_fin.isoformat() if horaire.date_fin else None,
                'actif': True,
            }
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


# ============================================================
# VUES SALAIRES
# ============================================================

def _get_mois_et_fin(request):
    mois_str = request.GET.get('mois')
    if mois_str:
        try:
            mois = date.fromisoformat(f"{mois_str}-01")
        except ValueError:
            mois = date.today().replace(day=1)
    else:
        mois = date.today().replace(day=1)
    fin_mois = (mois.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
    return mois, fin_mois


@login_required
@role_required(['patron'])
def salaires_calcul(request):
    mois, fin_mois = _get_mois_et_fin(request)
    institut_code = request.GET.get('institut', 'tous')
    vue = request.GET.get('vue', 'calcul')  # 'calcul', 'primes', 'avances'
    instituts = Institut.objects.all().order_by('nom')
    employes_all = Employe.objects.filter(actif=True).order_by('nom')
    types_prime = TypePrime.objects.filter(actif=True)

    def _filtre_institut(qs, champ_institut='employe__institut__code'):
        if institut_code == 'autres':
            return qs.filter(**{champ_institut.replace('__code', '__isnull'): True})
        elif institut_code != 'tous':
            return qs.filter(**{champ_institut: institut_code})
        return qs

    context = {
        'mois': mois,
        'instituts': instituts,
        'institut_code': institut_code,
        'vue': vue,
        'employes_all': employes_all,
        'types_prime': types_prime,
    }

    if vue == 'primes':
        primes = _filtre_institut(
            Prime.objects.filter(mois=mois).select_related('employe', 'employe__institut', 'type_prime')
        ).order_by('-date_creation')
        context['primes'] = primes
        context['total_primes'] = sum(p.montant for p in primes)

    elif vue == 'avances':
        statut_filtre = request.GET.get('statut', 'en_cours')
        avances_qs = Avance.objects.select_related('employe', 'employe__institut')
        if statut_filtre != 'tous':
            avances_qs = avances_qs.filter(statut=statut_filtre)
        avances_qs = _filtre_institut(avances_qs).order_by('-date')
        context['avances'] = avances_qs
        context['statut_filtre'] = statut_filtre
        context['total_en_cours'] = sum(
            a.reste_a_rembourser() for a in Avance.objects.filter(statut='en_cours')
        )

    else:  # vue == 'calcul'
        if institut_code == 'tous':
            employes = employes_all
        elif institut_code == 'autres':
            employes = Employe.objects.filter(institut__isnull=True, actif=True).order_by('nom')
        else:
            employes = Employe.objects.filter(institut__code=institut_code, actif=True).order_by('nom')

        calculs = []
        for emp in employes:
            calcul, created = CalculSalaire.objects.get_or_create(
                employe=emp, mois=mois,
                defaults={'salaire_base': emp.salaire_base}
            )
            if created or calcul.statut == 'brouillon':
                calcul.calculer()
            calculs.append(calcul)

        context['calculs'] = calculs
        context['total_net'] = sum(c.net_a_payer for c in calculs)
        context['total_base'] = sum(c.salaire_base for c in calculs)

    return render(request, 'gestion/salaires/calcul.html', context)


@login_required
@role_required(['patron'])
def primes_liste(request):
    mois, _ = _get_mois_et_fin(request)
    institut_code = request.GET.get('institut', 'tous')
    instituts = Institut.objects.all().order_by('nom')

    primes = Prime.objects.filter(mois=mois).select_related('employe', 'employe__institut', 'type_prime')
    if institut_code == 'autres':
        primes = primes.filter(employe__institut__isnull=True)
    elif institut_code != 'tous':
        primes = primes.filter(employe__institut__code=institut_code)

    employes = Employe.objects.filter(actif=True).order_by('nom')
    types_prime = TypePrime.objects.filter(actif=True)

    return render(request, 'gestion/salaires/primes.html', {
        'mois': mois,
        'primes': primes.order_by('-date_creation'),
        'employes': employes,
        'types_prime': types_prime,
        'instituts': instituts,
        'institut_code': institut_code,
        'total_primes': sum(p.montant for p in primes),
    })


@login_required
@role_required(['patron'])
def avances_liste(request):
    institut_code = request.GET.get('institut', 'tous')
    instituts = Institut.objects.all().order_by('nom')
    statut_filtre = request.GET.get('statut', 'en_cours')

    avances = Avance.objects.select_related('employe', 'employe__institut')
    if statut_filtre != 'tous':
        avances = avances.filter(statut=statut_filtre)
    if institut_code == 'autres':
        avances = avances.filter(employe__institut__isnull=True)
    elif institut_code != 'tous':
        avances = avances.filter(employe__institut__code=institut_code)

    employes = Employe.objects.filter(actif=True).order_by('nom')
    total_en_cours = sum(a.reste_a_rembourser() for a in Avance.objects.filter(statut='en_cours'))

    return render(request, 'gestion/salaires/avances.html', {
        'avances': avances.order_by('-date'),
        'employes': employes,
        'instituts': instituts,
        'institut_code': institut_code,
        'statut_filtre': statut_filtre,
        'total_en_cours': total_en_cours,
    })


@login_required
@role_required(['patron'])
@require_POST
def api_calculer_salaire(request):
    try:
        data = json.loads(request.body)
        employe = get_object_or_404(Employe, id=data['employe_id'])
        mois = date.fromisoformat(f"{data['mois']}-01")
        calcul, _ = CalculSalaire.objects.get_or_create(
            employe=employe, mois=mois,
            defaults={'salaire_base': employe.salaire_base}
        )
        net = calcul.calculer()
        return JsonResponse({
            'success': True, 'net_a_payer': net,
            'salaire_base': calcul.salaire_base,
            'total_primes': calcul.total_primes,
            'montant_retenue_absences': calcul.montant_retenue_absences,
            'total_avances_deduites': calcul.total_avances_deduites,
            'jours_travailles': calcul.jours_travailles,
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
@role_required(['patron'])
@require_POST
def api_valider_salaire(request, calcul_id):
    try:
        calcul = get_object_or_404(CalculSalaire, id=calcul_id)
        action = request.POST.get('action', 'valider')
        if action == 'valider' and calcul.statut == 'brouillon':
            calcul.statut = 'valide'
            calcul.valide_par = request.user.utilisateur
            calcul.date_validation = timezone.now()
        elif action == 'payer' and calcul.statut == 'valide':
            calcul.statut = 'paye'
            calcul.date_paiement = timezone.now()
            for detail in calcul.details_avances:
                try:
                    av = Avance.objects.get(employe=calcul.employe, date=date.fromisoformat(detail['date']), statut='en_cours')
                    av.montant_rembourse += detail['montant_deduit']
                    av.save()
                except Avance.DoesNotExist:
                    pass
        calcul.save()
        return JsonResponse({'success': True, 'statut': calcul.statut, 'message': f'Salaire {calcul.get_statut_display().lower()}'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
@role_required(['patron'])
@require_POST
def api_prime_creer(request):
    try:
        employe = get_object_or_404(Employe, id=request.POST.get('employe_id'))
        type_prime = get_object_or_404(TypePrime, id=request.POST.get('type_prime_id'))
        montant = int(request.POST.get('montant', 0))
        mois = date.fromisoformat(f"{request.POST.get('mois')}-01")
        if montant <= 0:
            return JsonResponse({'success': False, 'error': 'Montant invalide'}, status=400)
        Prime.objects.create(
            employe=employe, type_prime=type_prime, mois=mois, montant=montant,
            commentaire=request.POST.get('commentaire', '').strip(),
            cree_par=request.user.utilisateur,
        )
        CalculSalaire.objects.filter(employe=employe, mois=mois, statut='brouillon').delete()
        return JsonResponse({'success': True, 'message': f'Prime "{type_prime.nom}" de {montant:,} CFA ajoutée'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
@role_required(['patron'])
@require_POST
def api_prime_supprimer(request, prime_id):
    try:
        prime = get_object_or_404(Prime, id=prime_id)
        nom = f"{prime.type_prime.nom} — {prime.employe.get_full_name()}"
        CalculSalaire.objects.filter(employe=prime.employe, mois=prime.mois, statut='brouillon').delete()
        prime.delete()
        return JsonResponse({'success': True, 'message': f'Prime "{nom}" supprimée'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
@role_required(['patron'])
@require_POST
def api_avance_creer(request):
    try:
        employe = get_object_or_404(Employe, id=request.POST.get('employe_id'))
        montant = int(request.POST.get('montant', 0))
        nb_mois = int(request.POST.get('nombre_mois_remboursement', 1))
        date_avance = date.fromisoformat(request.POST.get('date'))
        if montant <= 0:
            return JsonResponse({'success': False, 'error': 'Montant invalide'}, status=400)
        avance = Avance.objects.create(
            employe=employe, date=date_avance, montant=montant,
            nombre_mois_remboursement=nb_mois,
            commentaire=request.POST.get('commentaire', '').strip(),
            cree_par=request.user.utilisateur,
        )
        return JsonResponse({'success': True, 'message': f'Avance de {montant:,} CFA — {avance.montant_mensuel:,} CFA/mois sur {nb_mois} mois'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
@role_required(['patron'])
@require_POST
def api_avance_supprimer(request, avance_id):
    try:
        avance = get_object_or_404(Avance, id=avance_id)
        if avance.montant_rembourse > 0:
            avance.statut = 'annule'
            avance.save()
            msg = f'Avance de {avance.employe.get_full_name()} annulée'
        else:
            nom = avance.employe.get_full_name()
            avance.delete()
            msg = f'Avance de {nom} supprimée'
        return JsonResponse({'success': True, 'message': msg})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


# ============================================================
# VUES STOCKS
# ============================================================

@login_required
@role_required(['patron', 'manager'])
def stocks_produits(request):
    """Liste des produits avec alertes de stock bas."""
    categorie_id = request.GET.get('categorie', 'tous')
    filtre_alerte = request.GET.get('alerte', '')

    produits = Produit.objects.filter(actif=True).select_related('categorie', 'unite', 'fournisseur')

    if categorie_id != 'tous' and categorie_id:
        produits = produits.filter(categorie_id=categorie_id)
    if filtre_alerte == '1':
        # Filtrer les produits en alerte (stock <= minimum)
        ids_alerte = [p.id for p in produits if p.en_alerte]
        produits = produits.filter(id__in=ids_alerte)

    nb_alertes = sum(1 for p in Produit.objects.filter(actif=True) if p.en_alerte)
    valeur_totale = sum(p.valeur_stock for p in produits)

    categories = CategorieProduit.objects.all()
    unites = UniteMesure.objects.all()
    fournisseurs = Fournisseur.objects.filter(actif=True)

    return render(request, 'gestion/stocks/produits.html', {
        'produits': produits,
        'categories': categories,
        'unites': unites,
        'fournisseurs': fournisseurs,
        'categorie_id': categorie_id,
        'filtre_alerte': filtre_alerte,
        'nb_alertes': nb_alertes,
        'valeur_totale': valeur_totale,
        'nb_produits': produits.count(),
    })


@login_required
@role_required(['patron', 'manager'])
def stocks_mouvements(request):
    """Historique des mouvements de stock."""
    from django.core.paginator import Paginator

    produit_id = request.GET.get('produit', 'tous')
    type_mouv = request.GET.get('type', 'tous')

    mouvements = MouvementStock.objects.select_related('produit', 'cree_par')

    if produit_id != 'tous' and produit_id:
        mouvements = mouvements.filter(produit_id=produit_id)
    if type_mouv != 'tous' and type_mouv:
        mouvements = mouvements.filter(type_mouvement=type_mouv)

    paginator = Paginator(mouvements, 50)
    page = request.GET.get('page', 1)
    mouvements_page = paginator.get_page(page)

    produits = Produit.objects.filter(actif=True).order_by('nom')

    return render(request, 'gestion/stocks/mouvements.html', {
        'mouvements': mouvements_page,
        'produits': produits,
        'produit_id': produit_id,
        'type_mouv': type_mouv,
    })


@login_required
@role_required(['patron'])
@require_POST
def api_produit_creer(request):
    try:
        nom = request.POST.get('nom', '').strip()
        if not nom:
            return JsonResponse({'success': False, 'error': 'Le nom est requis'}, status=400)

        categorie_id = request.POST.get('categorie_id', '').strip()
        unite_id = request.POST.get('unite_id', '').strip()
        fournisseur_id = request.POST.get('fournisseur_id', '').strip()

        produit = Produit.objects.create(
            nom=nom,
            reference=request.POST.get('reference', '').strip(),
            categorie=CategorieProduit.objects.filter(id=categorie_id).first() if categorie_id else None,
            unite=UniteMesure.objects.filter(id=unite_id).first() if unite_id else None,
            fournisseur=Fournisseur.objects.filter(id=fournisseur_id).first() if fournisseur_id else None,
            prix_achat=int(request.POST.get('prix_achat', 0) or 0),
            prix_vente=int(request.POST.get('prix_vente', 0) or 0),
            stock_actuel=int(request.POST.get('stock_actuel', 0) or 0),
            stock_minimum=int(request.POST.get('stock_minimum', 0) or 0),
        )
        return JsonResponse({'success': True, 'message': f'Produit "{produit.nom}" créé'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
@role_required(['patron'])
@require_POST
def api_produit_modifier(request, produit_id):
    try:
        produit = get_object_or_404(Produit, id=produit_id)
        nom = request.POST.get('nom', '').strip()
        if not nom:
            return JsonResponse({'success': False, 'error': 'Le nom est requis'}, status=400)

        categorie_id = request.POST.get('categorie_id', '').strip()
        unite_id = request.POST.get('unite_id', '').strip()
        fournisseur_id = request.POST.get('fournisseur_id', '').strip()

        produit.nom = nom
        produit.reference = request.POST.get('reference', '').strip()
        produit.categorie = CategorieProduit.objects.filter(id=categorie_id).first() if categorie_id else None
        produit.unite = UniteMesure.objects.filter(id=unite_id).first() if unite_id else None
        produit.fournisseur = Fournisseur.objects.filter(id=fournisseur_id).first() if fournisseur_id else None
        produit.prix_achat = int(request.POST.get('prix_achat', 0) or 0)
        produit.prix_vente = int(request.POST.get('prix_vente', 0) or 0)
        produit.stock_minimum = int(request.POST.get('stock_minimum', 0) or 0)
        produit.save()
        return JsonResponse({'success': True, 'message': f'Produit "{produit.nom}" modifié'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
@role_required(['patron'])
@require_POST
def api_produit_supprimer(request, produit_id):
    try:
        produit = get_object_or_404(Produit, id=produit_id)
        if produit.mouvements.exists():
            produit.actif = False
            produit.save()
            return JsonResponse({'success': True, 'message': f'Produit "{produit.nom}" désactivé (a des mouvements)'})
        nom = produit.nom
        produit.delete()
        return JsonResponse({'success': True, 'message': f'Produit "{nom}" supprimé'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
@role_required(['patron', 'manager'])
@require_POST
def api_mouvement_creer(request):
    try:
        produit_id = request.POST.get('produit_id', '').strip()
        if not produit_id:
            return JsonResponse({'success': False, 'error': 'Produit requis'}, status=400)

        produit = get_object_or_404(Produit, id=produit_id)
        type_mouv = request.POST.get('type_mouvement', '')
        quantite = int(request.POST.get('quantite', 0) or 0)
        if quantite <= 0:
            return JsonResponse({'success': False, 'error': 'La quantité doit être > 0'}, status=400)

        stock_avant = produit.stock_actuel

        if type_mouv == 'entree':
            produit.stock_actuel += quantite
        elif type_mouv in ('sortie', 'perte'):
            if quantite > produit.stock_actuel:
                return JsonResponse({'success': False, 'error': f'Stock insuffisant ({produit.stock_actuel} disponible)'}, status=400)
            produit.stock_actuel -= quantite
        elif type_mouv == 'inventaire':
            # Recalage : on fixe le stock à la valeur saisie
            produit.stock_actuel = quantite
            quantite = abs(quantite - stock_avant)  # diff pour le mouvement
        else:
            return JsonResponse({'success': False, 'error': 'Type de mouvement invalide'}, status=400)

        produit.save()

        MouvementStock.objects.create(
            produit=produit,
            type_mouvement=type_mouv,
            quantite=quantite,
            quantite_avant=stock_avant,
            quantite_apres=produit.stock_actuel,
            prix_unitaire=int(request.POST.get('prix_unitaire', produit.prix_achat) or produit.prix_achat),
            commentaire=request.POST.get('commentaire', '').strip() or None,
            cree_par=request.user.utilisateur if request.user.is_authenticated and hasattr(request.user, 'utilisateur') else None,
        )

        return JsonResponse({
            'success': True,
            'message': f'Mouvement enregistré — Stock {produit.nom} : {produit.stock_actuel}',
            'stock_actuel': produit.stock_actuel,
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


# ============================================================
# VUES STOCKS — PARAMÈTRES (catégories, unités, fournisseurs)
# ============================================================

@login_required
@role_required(['patron'])
def stocks_parametres(request):
    return render(request, 'gestion/stocks/parametres.html', {
        'categories': CategorieProduit.objects.annotate(nb=Count('produits')).order_by('nom'),
        'unites': UniteMesure.objects.annotate(nb=Count('produits')).order_by('nom'),
        'fournisseurs': Fournisseur.objects.annotate(nb=Count('produits')).order_by('nom'),
    })


@login_required
@role_required(['patron'])
@require_POST
def api_categorie_produit_creer(request):
    try:
        nom = request.POST.get('nom', '').strip()
        if not nom:
            return JsonResponse({'success': False, 'error': 'Nom requis'}, status=400)
        cat, created = CategorieProduit.objects.get_or_create(nom=nom, defaults={'description': request.POST.get('description', '').strip()})
        if not created:
            return JsonResponse({'success': False, 'error': 'Cette catégorie existe déjà'}, status=400)
        return JsonResponse({'success': True, 'message': f'Catégorie "{cat.nom}" créée', 'id': cat.id})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
@role_required(['patron'])
@require_POST
def api_categorie_produit_modifier(request, pk):
    try:
        cat = get_object_or_404(CategorieProduit, pk=pk)
        nom = request.POST.get('nom', '').strip()
        if not nom:
            return JsonResponse({'success': False, 'error': 'Nom requis'}, status=400)
        cat.nom = nom
        cat.description = request.POST.get('description', '').strip()
        cat.save()
        return JsonResponse({'success': True, 'message': f'Catégorie "{cat.nom}" modifiée'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
@role_required(['patron'])
@require_POST
def api_categorie_produit_supprimer(request, pk):
    try:
        cat = get_object_or_404(CategorieProduit, pk=pk)
        nb = cat.produits.count()
        if nb > 0:
            return JsonResponse({'success': False, 'error': f'Impossible : {nb} produit(s) utilisent cette catégorie'}, status=400)
        nom = cat.nom
        cat.delete()
        return JsonResponse({'success': True, 'message': f'Catégorie "{nom}" supprimée'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
@role_required(['patron'])
@require_POST
def api_unite_creer(request):
    try:
        nom = request.POST.get('nom', '').strip()
        if not nom:
            return JsonResponse({'success': False, 'error': 'Nom requis'}, status=400)
        unite, created = UniteMesure.objects.get_or_create(nom=nom, defaults={'abrv': request.POST.get('abrv', '').strip()})
        if not created:
            return JsonResponse({'success': False, 'error': 'Cette unité existe déjà'}, status=400)
        return JsonResponse({'success': True, 'message': f'Unité "{unite.nom}" créée', 'id': unite.id})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
@role_required(['patron'])
@require_POST
def api_unite_modifier(request, pk):
    try:
        unite = get_object_or_404(UniteMesure, pk=pk)
        nom = request.POST.get('nom', '').strip()
        if not nom:
            return JsonResponse({'success': False, 'error': 'Nom requis'}, status=400)
        unite.nom = nom
        unite.abrv = request.POST.get('abrv', '').strip()
        unite.save()
        return JsonResponse({'success': True, 'message': f'Unité "{unite.nom}" modifiée'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
@role_required(['patron'])
@require_POST
def api_unite_supprimer(request, pk):
    try:
        unite = get_object_or_404(UniteMesure, pk=pk)
        nb = unite.produits.count()
        if nb > 0:
            return JsonResponse({'success': False, 'error': f'Impossible : {nb} produit(s) utilisent cette unité'}, status=400)
        nom = unite.nom
        unite.delete()
        return JsonResponse({'success': True, 'message': f'Unité "{nom}" supprimée'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
@role_required(['patron'])
@require_POST
def api_fournisseur_creer(request):
    try:
        nom = request.POST.get('nom', '').strip()
        if not nom:
            return JsonResponse({'success': False, 'error': 'Nom requis'}, status=400)
        fo = Fournisseur.objects.create(
            nom=nom,
            telephone=request.POST.get('telephone', '').strip(),
            email=request.POST.get('email', '').strip(),
        )
        return JsonResponse({'success': True, 'message': f'Fournisseur "{fo.nom}" créé', 'id': fo.id})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
@role_required(['patron'])
@require_POST
def api_fournisseur_modifier(request, pk):
    try:
        fo = get_object_or_404(Fournisseur, pk=pk)
        nom = request.POST.get('nom', '').strip()
        if not nom:
            return JsonResponse({'success': False, 'error': 'Nom requis'}, status=400)
        fo.nom = nom
        fo.telephone = request.POST.get('telephone', '').strip()
        fo.email = request.POST.get('email', '').strip()
        fo.save()
        return JsonResponse({'success': True, 'message': f'Fournisseur "{fo.nom}" modifié'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
@role_required(['patron'])
@require_POST
def api_fournisseur_supprimer(request, pk):
    try:
        fo = get_object_or_404(Fournisseur, pk=pk)
        nb = fo.produits.count()
        if nb > 0:
            return JsonResponse({'success': False, 'error': f'Impossible : {nb} produit(s) liés à ce fournisseur'}, status=400)
        nom = fo.nom
        fo.delete()
        return JsonResponse({'success': True, 'message': f'Fournisseur "{nom}" supprimé'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


# ============================================================
# VUES DÉPENSES
# ============================================================

@login_required
@role_required(['patron', 'manager'])
def depenses_liste(request):
    from django.db.models import Sum

    mois_param = request.GET.get('mois', '')
    categorie_id = request.GET.get('categorie', 'tous')
    institut_code = request.GET.get('institut', 'tous')

    if mois_param:
        try:
            from datetime import datetime
            mois = datetime.strptime(mois_param, '%Y-%m').date()
        except ValueError:
            mois = date.today().replace(day=1)
    else:
        mois = date.today().replace(day=1)

    fin_mois = (mois.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)

    depenses = Depense.objects.filter(date__range=[mois, fin_mois]).select_related('categorie', 'institut', 'cree_par')

    if categorie_id != 'tous' and categorie_id:
        depenses = depenses.filter(categorie_id=categorie_id)
    if institut_code != 'tous':
        if institut_code == 'autres':
            depenses = depenses.filter(institut__isnull=True)
        else:
            depenses = depenses.filter(institut__code=institut_code)

    total = depenses.aggregate(t=Sum('montant'))['t'] or 0

    # Totaux par catégorie
    par_categorie = {}
    for d in depenses:
        nom_cat = d.categorie.nom
        par_categorie[nom_cat] = par_categorie.get(nom_cat, 0) + d.montant

    categories = CategoriDepense.objects.filter(type__in=['informelle', 'les_deux'])
    instituts = Institut.objects.all()

    return render(request, 'gestion/depenses/liste.html', {
        'depenses': depenses,
        'mois': mois,
        'total': total,
        'par_categorie': sorted(par_categorie.items(), key=lambda x: -x[1]),
        'categories': categories,
        'instituts': instituts,
        'categorie_id': categorie_id,
        'institut_code': institut_code,
    })


@login_required
@role_required(['patron', 'manager'])
@require_POST
def api_depense_creer(request):
    try:
        categorie_id = request.POST.get('categorie_id', '').strip()
        montant = int(request.POST.get('montant', 0) or 0)
        date_str = request.POST.get('date', '').strip()

        if not categorie_id:
            return JsonResponse({'success': False, 'error': 'Catégorie requise'}, status=400)
        if montant <= 0:
            return JsonResponse({'success': False, 'error': 'Montant invalide'}, status=400)
        if not date_str:
            return JsonResponse({'success': False, 'error': 'Date requise'}, status=400)

        from datetime import datetime
        date_depense = datetime.strptime(date_str, '%Y-%m-%d').date()

        institut_code = request.POST.get('institut_code', '').strip()
        institut = Institut.objects.filter(code=institut_code).first() if institut_code else None

        depense = Depense.objects.create(
            categorie=get_object_or_404(CategoriDepense, id=categorie_id),
            institut=institut,
            montant=montant,
            date=date_depense,
            description=request.POST.get('description', '').strip() or None,
            beneficiaire=request.POST.get('beneficiaire', '').strip(),
            mode_paiement=request.POST.get('mode_paiement', 'especes'),
            cree_par=request.user.utilisateur if request.user.is_authenticated and hasattr(request.user, 'utilisateur') else None,
        )
        return JsonResponse({'success': True, 'message': f'Dépense de {depense.montant} CFA enregistrée'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
@role_required(['patron'])
@require_POST
def api_depense_supprimer(request, depense_id):
    try:
        depense = get_object_or_404(Depense, id=depense_id)
        montant = depense.montant
        cat = depense.categorie.nom
        depense.delete()
        return JsonResponse({'success': True, 'message': f'Dépense {cat} — {montant} CFA supprimée'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
@role_required(['patron'])
@require_POST
def api_categorie_depense_creer(request):
    try:
        nom = request.POST.get('nom', '').strip()
        if not nom:
            return JsonResponse({'success': False, 'error': 'Nom requis'}, status=400)
        type_cat = request.POST.get('type', 'les_deux')
        if type_cat not in ['informelle', 'recurrente', 'les_deux']:
            type_cat = 'les_deux'
        cat, created = CategoriDepense.objects.get_or_create(
            nom=nom,
            defaults={'description': request.POST.get('description', '').strip(), 'type': type_cat},
        )
        if not created:
            return JsonResponse({'success': False, 'error': 'Cette catégorie existe déjà'}, status=400)
        return JsonResponse({'success': True, 'message': f'Catégorie "{cat.nom}" créée', 'id': cat.id, 'nom': cat.nom, 'type': cat.type})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
@role_required(['patron'])
@require_POST
def api_categorie_depense_modifier(request, pk):
    try:
        cat = get_object_or_404(CategoriDepense, pk=pk)
        nom = request.POST.get('nom', '').strip()
        if not nom:
            return JsonResponse({'success': False, 'error': 'Nom requis'}, status=400)
        if CategoriDepense.objects.filter(nom=nom).exclude(pk=pk).exists():
            return JsonResponse({'success': False, 'error': 'Ce nom est déjà utilisé'}, status=400)
        type_cat = request.POST.get('type', 'les_deux')
        if type_cat not in ['informelle', 'recurrente', 'les_deux']:
            type_cat = 'les_deux'
        cat.nom = nom
        cat.type = type_cat
        cat.save(update_fields=['nom', 'type'])
        return JsonResponse({'success': True, 'message': f'Catégorie modifiée', 'id': cat.id, 'nom': cat.nom, 'type': cat.type})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
@role_required(['patron'])
@require_POST
def api_categorie_depense_supprimer(request, pk):
    try:
        cat = get_object_or_404(CategoriDepense, pk=pk)
        nb = cat.depenses.count()
        if nb > 0:
            return JsonResponse({'success': False, 'error': f'Impossible : {nb} dépense(s) utilisent cette catégorie'}, status=400)
        nom = cat.nom
        cat.delete()
        return JsonResponse({'success': True, 'message': f'Catégorie "{nom}" supprimée'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


# ─────────────────────────── BILAN ───────────────────────────

@login_required
@role_required(['patron'])
def bilan_mensuel(request):
    from django.db.models import Sum

    mois_param = request.GET.get('mois', '')
    institut_code = request.GET.get('institut', 'tous')

    if mois_param:
        try:
            from datetime import datetime
            mois = datetime.strptime(mois_param, '%Y-%m').date()
        except ValueError:
            mois = date.today().replace(day=1)
    else:
        mois = date.today().replace(day=1)

    fin_mois = (mois.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)

    instituts = Institut.objects.all().order_by('nom')

    # ── Recettes (Paiements hors forfait/offert) ──
    paiements_qs = Paiement.objects.filter(
        date__date__gte=mois,
        date__date__lte=fin_mois,
    ).exclude(mode__in=['forfait', 'offert'])

    if institut_code != 'tous':
        paiements_qs = paiements_qs.filter(
            rendez_vous__employe__institut__code=institut_code
        )

    ca_rdv = paiements_qs.aggregate(s=Sum('montant'))['s'] or 0

    # Recettes par mode de paiement
    recettes_par_mode = {}
    for p in paiements_qs.values('mode').annotate(total=Sum('montant')).order_by('-total'):
        recettes_par_mode[p['mode']] = p['total']

    # Ventes de produits
    ventes_produits_qs = VenteProduit.objects.filter(
        date__date__gte=mois,
        date__date__lte=fin_mois,
    )
    if institut_code != 'tous':
        ventes_produits_qs = ventes_produits_qs.filter(institut__code=institut_code)
    ca_ventes_produits = ventes_produits_qs.aggregate(s=Sum('montant_total'))['s'] or 0

    total_recettes = ca_rdv + ca_ventes_produits

    # ── Dépenses ──
    depenses_qs = Depense.objects.filter(date__gte=mois, date__lte=fin_mois)
    if institut_code != 'tous':
        depenses_qs = depenses_qs.filter(
            Q(institut__code=institut_code) | Q(institut__isnull=True)
        ) if institut_code else depenses_qs

    total_depenses = depenses_qs.aggregate(s=Sum('montant'))['s'] or 0

    # Dépenses par catégorie
    depenses_par_cat = list(
        depenses_qs.values('categorie__nom').annotate(total=Sum('montant')).order_by('-total')
    )

    # ── Bénéfice net ──
    benefice = total_recettes - total_depenses
    taux_charges = round((total_depenses / total_recettes * 100), 1) if total_recettes else 0

    # ── Évolution N derniers mois ──
    try:
        nb_mois = int(request.GET.get('nb_mois', 6))
        if nb_mois not in (3, 6, 12):
            nb_mois = 6
    except (ValueError, TypeError):
        nb_mois = 6

    evolution = []
    for i in range(nb_mois - 1, -1, -1):
        month = mois.month - i
        year = mois.year
        while month <= 0:
            month += 12
            year -= 1
        m_debut = date(year, month, 1)
        m_fin = (m_debut.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)

        rec_qs = Paiement.objects.filter(
            date__date__gte=m_debut,
            date__date__lte=m_fin,
        ).exclude(mode__in=['forfait', 'offert'])
        if institut_code != 'tous':
            rec_qs = rec_qs.filter(rendez_vous__employe__institut__code=institut_code)
        rec = rec_qs.aggregate(s=Sum('montant'))['s'] or 0

        vp_qs = VenteProduit.objects.filter(date__date__gte=m_debut, date__date__lte=m_fin)
        if institut_code != 'tous':
            vp_qs = vp_qs.filter(institut__code=institut_code)
        rec += vp_qs.aggregate(s=Sum('montant_total'))['s'] or 0

        dep_qs = Depense.objects.filter(date__gte=m_debut, date__lte=m_fin)
        if institut_code != 'tous':
            dep_qs = dep_qs.filter(Q(institut__code=institut_code) | Q(institut__isnull=True))
        dep = dep_qs.aggregate(s=Sum('montant'))['s'] or 0

        evolution.append({
            'label': m_debut.strftime('%b %Y'),
            'recettes': rec,
            'depenses': dep,
            'benefice': rec - dep,
        })

    import json as _json
    evolution_json = _json.dumps(evolution)

    return render(request, 'gestion/bilan/mensuel.html', {
        'mois': mois,
        'fin_mois': fin_mois,
        'instituts': instituts,
        'institut_code': institut_code,
        'nb_mois': nb_mois,
        'total_recettes': total_recettes,
        'ca_rdv': ca_rdv,
        'ca_ventes_produits': ca_ventes_produits,
        'total_depenses': total_depenses,
        'benefice': benefice,
        'taux_charges': taux_charges,
        'recettes_par_mode': recettes_par_mode,
        'depenses_par_cat': depenses_par_cat,
        'evolution_json': evolution_json,
    })


# ─────────────────────────── VENTES PRODUITS ───────────────────────────

@login_required
@role_required(['patron', 'manager'])
def ventes_caisse(request):
    from django.db.models import Sum
    utilisateur = request.user.utilisateur
    instituts = Institut.objects.all().order_by('nom')

    if utilisateur.role == 'patron':
        institut_code = request.GET.get('institut', instituts.first().code if instituts.exists() else 'palais')
    else:
        institut_code = utilisateur.institut.code if utilisateur.institut else 'palais'

    institut = Institut.objects.filter(code=institut_code).first()
    produits = Produit.objects.filter(actif=True, prix_vente__gt=0).select_related('unite', 'categorie').order_by('categorie__nom', 'nom')
    clients = Client.objects.filter(actif=True).order_by('nom')

    today = date.today()
    ventes_qs = VenteProduit.objects.filter(institut=institut, date__date=today).select_related('client', 'effectue_par').prefetch_related('lignes__produit').order_by('-date')
    agg = ventes_qs.aggregate(s=Sum('montant_total'))
    ventes_jour = agg['s'] or 0
    nb_ventes = ventes_qs.count()
    total_especes = ventes_qs.filter(mode_paiement='especes').aggregate(s=Sum('montant_total'))['s'] or 0
    total_carte = ventes_qs.filter(mode_paiement='carte').aggregate(s=Sum('montant_total'))['s'] or 0

    return render(request, 'gestion/ventes/caisse.html', {
        'instituts': instituts,
        'institut': institut,
        'institut_code': institut_code,
        'produits': produits,
        'clients': clients,
        'ventes_jour': ventes_jour,
        'nb_ventes': nb_ventes,
        'total_especes': total_especes,
        'total_carte': total_carte,
        'ventes_du_jour': ventes_qs,
        'can_edit': utilisateur.role == 'patron',
    })


@login_required
@role_required(['patron', 'manager'])
def ventes_historique(request):
    from django.db.models import Sum
    utilisateur = request.user.utilisateur
    instituts = Institut.objects.all().order_by('nom')

    mois_param = request.GET.get('mois', '')
    if mois_param:
        try:
            from datetime import datetime
            mois = datetime.strptime(mois_param, '%Y-%m').date()
        except ValueError:
            mois = date.today().replace(day=1)
    else:
        mois = date.today().replace(day=1)

    fin_mois = (mois.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)

    if utilisateur.role == 'patron':
        institut_code = request.GET.get('institut', 'tous')
    else:
        institut_code = utilisateur.institut.code if utilisateur.institut else 'tous'

    ventes = VenteProduit.objects.filter(
        date__date__gte=mois, date__date__lte=fin_mois
    ).select_related('institut', 'client', 'effectue_par').prefetch_related('lignes__produit')

    if institut_code != 'tous':
        ventes = ventes.filter(institut__code=institut_code)

    total = ventes.aggregate(s=Sum('montant_total'))['s'] or 0

    from django.core.paginator import Paginator
    paginator = Paginator(ventes, 50)
    page = request.GET.get('page', 1)
    ventes_page = paginator.get_page(page)

    return render(request, 'gestion/ventes/historique.html', {
        'instituts': instituts,
        'institut_code': institut_code,
        'mois': mois,
        'ventes': ventes_page,
        'total': total,
    })


@login_required
@role_required(['patron', 'manager'])
@require_POST
def api_vendre(request):
    try:
        from core.models import CarteCadeau, UtilisationCarteCadeau
        data = json.loads(request.body)
        utilisateur = request.user.utilisateur

        # Client obligatoire
        client_id = data.get('client_id')
        if not client_id:
            return JsonResponse({'success': False, 'error': 'Le client est obligatoire.'}, status=400)
        client = get_object_or_404(Client, id=client_id)

        institut = get_object_or_404(Institut, id=data['institut_id'])

        # Paiement
        mode_paiement   = data.get('mode_paiement', 'especes')
        mode_paiement_2 = data.get('mode_paiement_2') or None
        montant_paiement_1 = int(data.get('montant_paiement_1', 0) or 0)

        # Carte cadeau
        carte_id       = data.get('carte_id')
        montant_carte  = int(data.get('montant_carte', 0) or 0)
        carte = None
        if carte_id and montant_carte > 0:
            carte = CarteCadeau.objects.filter(id=carte_id, statut='active', solde__gte=montant_carte).first()
            if not carte:
                return JsonResponse({'success': False, 'error': 'Carte cadeau invalide ou solde insuffisant.'}, status=400)

        # Créer la vente
        vente = VenteProduit.objects.create(
            institut=institut,
            client=client,
            mode_paiement=mode_paiement,
            mode_paiement_2=mode_paiement_2,
            montant_paiement_1=montant_paiement_1,
            carte_cadeau_utilisee=carte,
            montant_carte_utilise=montant_carte,
            effectue_par=utilisateur,
        )

        # Lignes + mouvements stock
        for item in data['items']:
            produit = get_object_or_404(Produit, id=item['produit_id'])
            quantite = int(item['quantite'])
            if produit.stock_actuel < quantite:
                vente.delete()
                return JsonResponse({'success': False, 'error': f'Stock insuffisant pour {produit.nom} (stock : {produit.stock_actuel})'}, status=400)

            LigneVenteProduit.objects.create(
                vente=vente,
                produit=produit,
                quantite=quantite,
                prix_unitaire=produit.prix_vente,
            )
            stock_avant = produit.stock_actuel
            produit.stock_actuel -= quantite
            produit.save(update_fields=['stock_actuel'])
            MouvementStock.objects.create(
                produit=produit,
                type_mouvement='sortie',
                quantite=quantite,
                quantite_avant=stock_avant,
                quantite_apres=produit.stock_actuel,
                prix_unitaire=produit.prix_vente,
                commentaire='Vente produit',
                cree_par=utilisateur,
            )

        vente.calculer_total()

        # Déduire la carte cadeau
        if carte and montant_carte > 0:
            carte.solde -= montant_carte
            if carte.solde <= 0:
                carte.solde = 0
                carte.statut = 'soldee'
            carte.date_derniere_utilisation = timezone.now()
            carte.save()
            UtilisationCarteCadeau.objects.create(
                carte=carte,
                montant=montant_carte,
                institut=institut,
                enregistre_par=utilisateur,
            )

        return JsonResponse({
            'success': True,
            'message': f'Vente enregistrée — {vente.montant_total:,} CFA',
            'vente_id': vente.id,
            'total': vente.montant_total,
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


# ─────────────────────────── RÉCONCILIATION ───────────────────────────

@login_required
@role_required(['patron'])
def reconciliation_index(request):
    instituts = Institut.objects.all().order_by('nom')
    institut_code = request.GET.get('institut', instituts.first().code if instituts.exists() else 'palais')
    institut = Institut.objects.filter(code=institut_code).first()

    date_param = request.GET.get('date', '')
    try:
        from datetime import datetime as dt
        jour = dt.strptime(date_param, '%Y-%m-%d').date() if date_param else date.today()
    except ValueError:
        jour = date.today()

    reconciliation, created = ReconciliationCaisse.objects.get_or_create(
        date=jour, institut=institut,
    )
    if created or not reconciliation.valide:
        reconciliation.calculer()

    historique = ReconciliationCaisse.objects.filter(institut=institut).order_by('-date')[:30]

    return render(request, 'gestion/reconciliation/index.html', {
        'instituts': instituts,
        'institut': institut,
        'institut_code': institut_code,
        'jour': jour,
        'reconciliation': reconciliation,
        'historique': historique,
    })


@login_required
@role_required(['patron'])
@require_POST
def api_reconciliation_calculer(request):
    try:
        data = json.loads(request.body)
        institut = get_object_or_404(Institut, id=data['institut_id'])
        from datetime import datetime as dt
        jour = dt.strptime(data['date'], '%Y-%m-%d').date()

        rec, _ = ReconciliationCaisse.objects.get_or_create(date=jour, institut=institut)
        rec.caisse_especes = int(data.get('especes', 0))
        rec.caisse_wave = int(data.get('wave', 0))
        rec.caisse_om = int(data.get('om', 0))
        rec.caisse_carte = int(data.get('carte', 0))
        rec.commentaire_ecart = data.get('commentaire', '')
        rec.save()
        rec.calculer()

        return JsonResponse({
            'success': True,
            'recettes_rdv': rec.recettes_rdv,
            'recettes_ventes': rec.recettes_ventes,
            'recettes_attendues': rec.recettes_attendues,
            'caisse_reelle': rec.caisse_reelle,
            'ecart': rec.ecart,
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
@role_required(['patron'])
@require_POST
def api_reconciliation_valider(request):
    try:
        data = json.loads(request.body)
        institut = get_object_or_404(Institut, id=data['institut_id'])
        from datetime import datetime as dt
        jour = dt.strptime(data['date'], '%Y-%m-%d').date()

        rec = get_object_or_404(ReconciliationCaisse, date=jour, institut=institut)
        rec.valide = True
        rec.valide_par = request.user.utilisateur
        rec.date_validation = timezone.now()
        rec.save()

        return JsonResponse({'success': True, 'message': 'Réconciliation validée'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


# =====================================================================
# INVENTAIRE
# =====================================================================

@login_required
@role_required(['patron', 'manager'])
def inventaire_liste(request):
    user = request.user.utilisateur
    institut = user.institut
    inventaires = Inventaire.objects.filter(effectue_par__institut=institut).select_related('effectue_par')[:50]
    return render(request, 'gestion/stocks/inventaire_liste.html', {
        'inventaires': inventaires,
        'titre': 'Inventaires',
    })


@login_required
@role_required(['patron', 'manager'])
def inventaire_nouveau(request):
    user = request.user.utilisateur
    institut = user.institut
    produits = Produit.objects.filter(actif=True).select_related('categorie').order_by('categorie__nom', 'nom')

    inv = Inventaire.objects.create(
        date=timezone.now().date(),
        effectue_par=user,
    )
    for p in produits:
        LigneInventaire.objects.create(
            inventaire=inv,
            produit=p,
            quantite_theorique=p.stock_actuel,
            quantite_reelle=p.stock_actuel,
            ecart=0,
        )
    return redirect('gestion:inventaire_detail', inventaire_id=inv.id)


@login_required
@role_required(['patron', 'manager'])
def inventaire_detail(request, inventaire_id):
    user = request.user.utilisateur
    inv = get_object_or_404(Inventaire, id=inventaire_id)
    lignes = inv.lignes.select_related('produit', 'produit__categorie').order_by('produit__categorie__nom', 'produit__nom')

    # Regrouper par catégorie
    categories = {}
    for ligne in lignes:
        cat_nom = ligne.produit.categorie.nom if ligne.produit.categorie else 'Sans catégorie'
        if cat_nom not in categories:
            categories[cat_nom] = []
        categories[cat_nom].append(ligne)

    stat_total = lignes.count()
    stat_ecarts = lignes.exclude(ecart=0).count()

    return render(request, 'gestion/stocks/inventaire_detail.html', {
        'inv': inv,
        'categories': categories,
        'titre': f'Inventaire du {inv.date.strftime("%d/%m/%Y")}',
        'stat_total': stat_total,
        'stat_ecarts': stat_ecarts,
    })


@login_required
@role_required(['patron', 'manager'])
@require_POST
def api_inventaire_saisir(request, inventaire_id, ligne_id):
    try:
        inv = get_object_or_404(Inventaire, id=inventaire_id)
        if inv.statut != 'en_cours':
            return JsonResponse({'success': False, 'error': 'Inventaire clôturé'}, status=400)

        data = json.loads(request.body)
        ligne = get_object_or_404(LigneInventaire, id=ligne_id, inventaire=inv)
        ligne.quantite_reelle = int(data.get('quantite_reelle', 0))
        ligne.ecart = ligne.quantite_reelle - ligne.quantite_theorique
        ligne.commentaire = data.get('commentaire', '')
        ligne.save()

        return JsonResponse({
            'success': True,
            'ecart': ligne.ecart,
            'quantite_reelle': ligne.quantite_reelle,
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
@role_required(['patron'])
@require_POST
def api_inventaire_cloturer(request, inventaire_id):
    try:
        user = request.user.utilisateur
        inv = get_object_or_404(Inventaire, id=inventaire_id)
        if inv.statut != 'en_cours':
            return JsonResponse({'success': False, 'error': 'Inventaire déjà clôturé'}, status=400)

        # Créer des mouvements stock pour corriger les écarts
        for ligne in inv.lignes.select_related('produit'):
            if ligne.ecart != 0:
                stock_avant = ligne.produit.stock_actuel
                MouvementStock.objects.create(
                    produit=ligne.produit,
                    type_mouvement='inventaire',
                    quantite=ligne.ecart,
                    quantite_avant=stock_avant,
                    quantite_apres=ligne.quantite_reelle,
                    commentaire=f'Ajustement inventaire #{inv.id}',
                    cree_par=user,
                )
                ligne.produit.stock_actuel = ligne.quantite_reelle
                ligne.produit.save(update_fields=['stock_actuel'])

        inv.statut = 'termine'
        inv.date_cloture = timezone.now()
        inv.save()

        return JsonResponse({'success': True, 'message': 'Inventaire clôturé'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


# =====================================================================
# DÉPENSES RÉCURRENTES
# =====================================================================

@login_required
@role_required(['patron', 'manager'])
def depenses_recurrentes(request):
    user = request.user.utilisateur
    institut = user.institut

    from datetime import date
    mois_courant = date.today().replace(day=1)

    recurrentes = DepenseRecurrente.objects.filter(institut=institut, actif=True).select_related('categorie')

    # Générer les validations manquantes pour le mois en cours
    for dr in recurrentes:
        ValidationDepenseRecurrente.objects.get_or_create(
            depense_recurrente=dr,
            mois=mois_courant,
            defaults={'statut': 'en_attente'},
        )

    validations_en_attente = ValidationDepenseRecurrente.objects.filter(
        depense_recurrente__institut=institut,
        mois=mois_courant,
        statut='en_attente',
    ).select_related('depense_recurrente', 'depense_recurrente__categorie')

    categories = CategoriDepense.objects.filter(type__in=['recurrente', 'les_deux']).order_by('nom')

    MODES_PAIEMENT = [
        ('especes', 'Espèces'), ('virement', 'Virement'), ('cheque', 'Chèque'),
        ('carte', 'Carte bancaire'), ('om', 'Orange Money'), ('wave', 'Wave'), ('autre', 'Autre'),
    ]

    return render(request, 'gestion/depenses/recurrentes.html', {
        'validations_en_attente': validations_en_attente,
        'mois_courant': mois_courant,
        'categories': categories,
        'MODES_PAIEMENT': MODES_PAIEMENT,
    })


@login_required
@role_required(['patron', 'manager'])
def depenses_parametres(request):
    user = request.user.utilisateur
    institut = user.institut

    recurrentes = DepenseRecurrente.objects.filter(
        institut=institut
    ).select_related('categorie').order_by('nom')

    categories = CategoriDepense.objects.all().order_by('nom')

    MODES_PAIEMENT = [
        ('especes', 'Espèces'), ('virement', 'Virement'), ('cheque', 'Chèque'),
        ('carte', 'Carte bancaire'), ('om', 'Orange Money'), ('wave', 'Wave'), ('autre', 'Autre'),
    ]

    return render(request, 'gestion/depenses/parametres.html', {
        'recurrentes': recurrentes,
        'categories': categories,
        'MODES_PAIEMENT': MODES_PAIEMENT,
    })


@login_required
@role_required(['patron', 'manager'])
@require_POST
def api_depense_recurrente_creer(request):
    try:
        user = request.user.utilisateur
        data = json.loads(request.body)
        categorie = get_object_or_404(CategoriDepense, id=data['categorie_id'])
        dr = DepenseRecurrente.objects.create(
            nom=data['nom'].strip(),
            categorie=categorie,
            institut=user.institut,
            montant=int(data['montant']),
            beneficiaire=data.get('beneficiaire', '').strip(),
            mode_paiement=data.get('mode_paiement', 'especes'),
            frequence=data.get('frequence', 'mensuel'),
            jour_du_mois=int(data.get('jour_du_mois', 1)),
            actif=True,
            cree_par=user,
        )
        return JsonResponse({'success': True, 'id': dr.id, 'nom': dr.nom})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
@role_required(['patron', 'manager'])
@require_POST
def api_depense_recurrente_modifier(request, dr_id):
    try:
        data = json.loads(request.body)
        dr = get_object_or_404(DepenseRecurrente, id=dr_id)
        if 'categorie_id' in data:
            dr.categorie = get_object_or_404(CategoriDepense, id=data['categorie_id'])
        if 'nom' in data:
            dr.nom = data['nom'].strip()
        if 'montant' in data:
            dr.montant = int(data['montant'])
        if 'beneficiaire' in data:
            dr.beneficiaire = data['beneficiaire'].strip()
        if 'mode_paiement' in data:
            dr.mode_paiement = data['mode_paiement']
        if 'frequence' in data:
            dr.frequence = data['frequence']
        if 'jour_du_mois' in data:
            dr.jour_du_mois = int(data['jour_du_mois'])
        if 'actif' in data:
            dr.actif = bool(data['actif'])
        dr.save()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
@role_required(['patron'])
@require_POST
def api_depense_recurrente_supprimer(request, dr_id):
    try:
        dr = get_object_or_404(DepenseRecurrente, id=dr_id)
        dr.delete()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
@role_required(['patron', 'manager'])
@require_POST
def api_validation_valider(request, val_id):
    try:
        user = request.user.utilisateur
        val = get_object_or_404(ValidationDepenseRecurrente, id=val_id)
        dr = val.depense_recurrente

        # Créer la dépense réelle
        depense = Depense.objects.create(
            categorie=dr.categorie,
            institut=dr.institut,
            montant=dr.montant,
            description=dr.nom,
            beneficiaire=dr.beneficiaire,
            mode_paiement=dr.mode_paiement,
            date=val.mois,
            cree_par=user,
        )
        val.statut = 'valide'
        val.depense = depense
        val.date_validation = timezone.now()
        val.valide_par = user
        val.save()

        return JsonResponse({'success': True, 'message': 'Dépense créée et validée'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
@role_required(['patron', 'manager'])
@require_POST
def api_validation_ignorer(request, val_id):
    try:
        user = request.user.utilisateur
        val = get_object_or_404(ValidationDepenseRecurrente, id=val_id)
        val.statut = 'ignore'
        val.date_validation = timezone.now()
        val.valide_par = user
        val.save()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)
